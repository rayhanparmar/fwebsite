from pathlib import Path
from dotenv import load_dotenv
from fastapi import Form, Query
from whatsapp_service import send_flow, send_text_message, send_document
from flow_crypto import decrypt_request, encrypt_response
from cloudinary.uploader import destroy
from io import BytesIO
import csv
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import logging
import bcrypt
import re
import jwt
import uuid
import threading
import random
import requests as http_requests
import boto3
from botocore.exceptions import ClientError
from fastapi import FastAPI, APIRouter, HTTPException, Request, UploadFile, File, Form
from fastapi.responses import Response, PlainTextResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field
from typing import List, Optional, Dict
from datetime import datetime, timezone, timedelta
from bson import ObjectId
from cloudinary_service import (
    upload_whatsapp_image,
    upload_whatsapp_video,
)
from pdf_service import create_order_pdf

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak
)

AWS_BUCKET_NAME = os.getenv("AWS_BUCKET_NAME")
AWS_REGION = os.getenv("AWS_REGION")

s3 = boto3.client(
    "s3",
    region_name=AWS_REGION,
    aws_access_key_id=os.getenv("AWS_ACCESS_KEY_ID"),
    aws_secret_access_key=os.getenv("AWS_SECRET_ACCESS_KEY"),
)

def upload_to_s3(file_data, filename, content_type):

    s3.upload_fileobj(
        file_data,
        AWS_BUCKET_NAME,
        filename,
        ExtraArgs={
            "ContentType": content_type
        }
    )

    return f"https://{AWS_BUCKET_NAME}.s3.{AWS_REGION}.amazonaws.com/{filename}"

# MongoDB
mongo_url = os.environ['MONGO_URL']
# client = AsyncIOMotorClient(mongo_url)
client = AsyncIOMotorClient(
    mongo_url,
    tls=True,
    tlsAllowInvalidCertificates=True,
    serverSelectionTimeoutMS=30000
)
db = client[os.environ['DB_NAME']]
whatsapp_db = client["whatsapp_orders"]
whatsapp_orders = whatsapp_db["orders"]
counters = whatsapp_db["counters"]

JWT_SECRET = os.environ.get('JWT_SECRET', 'fallback-secret-change-me')
JWT_ALGORITHM = "HS256"

pending_video_uploads = {}

video_waiting_users = {}
video_upload_timers = {}

VIDEO_TIMEOUT = 600


async def get_next_order_id():

    result = await counters.find_one_and_update(
        {"_id": "order_counter"},
        {"$inc": {"sequence": 1}},
        upsert=True,
        return_document=True
    )

    sequence = result["sequence"]

    return f"ORD-{sequence:06d}"
def video_upload_timeout(sender):

    if sender not in video_waiting_users:
        return

    pending_video_uploads.pop(sender, None)
    video_waiting_users.pop(sender, None)
    video_upload_timers.pop(sender, None)

    send_text_message(
        sender,
        """⌛ Video upload time expired.

Your order has been submitted successfully without a reference video.

Thank you!"""
    )

    print(f"Video upload timed out for {sender}")

app = FastAPI()
print("VERIFY_TOKEN =", os.getenv("VERIFY_TOKEN"))
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ──── CONSTANTS ────
CATEGORIES = [
    {"name": "Bali", "prefix": "BL", "slug": "bali"},
    {"name": "Bangle/Kada", "prefix": "BK", "slug": "bangle-kada"},
    {"name": "Bracelet", "prefix": "BR", "slug": "bracelet"},
    {"name": "Chain + Multilayer", "prefix": "CM", "slug": "chain-multilayer"},
    {"name": "Cufflink", "prefix": "CF", "slug": "cufflink"},
    {"name": "Brooch", "prefix": "BC", "slug": "brooch"},
    {"name": "Earring", "prefix": "ER", "slug": "earring"},
    {"name": "Haathpaan", "prefix": "HP", "slug": "haathpaan"},
    {"name": "Maang Tikka", "prefix": "MT", "slug": "maang-tikka"},
    {"name": "Mangal Sutra", "prefix": "MS", "slug": "mangal-sutra"},
    {"name": "Necklace", "prefix": "NK", "slug": "necklace"},
    {"name": "Nose Pin", "prefix": "NP", "slug": "nose-pin"},
    {"name": "Pendant + Dancing Stone", "prefix": "PD", "slug": "pendant-dancing-stone"},
    {"name": "Ring + Titanium Ring", "prefix": "RT", "slug": "ring-titanium-ring"},
    {"name": "Tops", "prefix": "TP", "slug": "tops"},
    {"name": "Watch Belt", "prefix": "WB", "slug": "watch-belt"},
    {"name": "Full Set", "prefix": "FS", "slug": "full-set"},
]

CATEGORY_ALIASES = {
    "Bali": ["Bali"],

    "Bangle/Kada": ["Bangle", "Kada"],

    "Bracelet": ["Bracelet"],

    "Chain + Multilayer": ["Chains"],

    "Cufflink": ["Cufflinks"],

    "Brooch": ["Brooch"],

    "Earring": ["Earrings"],

    "Haathpaan": ["Hath Pan"],

    "Maang Tikka": ["Maang Tikka"],

    "Mangal Sutra": ["Mangal Sutra"],

    "Necklace": ["Necklace"],

    "Nose Pin": ["Nose Pin"],

    "Pendant + Dancing Stone": ["Pendant"],

    "Ring + Titanium Ring": ["Rings"],

    "Tops": ["Tops"],

    "Watch Belt": ["Watchbelts"],

    "Full Set": ["Full Set"],
}

STOCK_IMAGES = {
    "Bali": ["https://images.unsplash.com/photo-1723361656146-f201d215c49c?w=500&q=80"],
    "Bangle": ["https://images.unsplash.com/photo-1723361656145-b481be3f9e05?w=500&q=80"],
    "Kada": ["https://images.unsplash.com/photo-1723361656062-ed14986c7f1a?w=500&q=80"],
    "Bracelet": ["https://images.unsplash.com/photo-1723361656145-b481be3f9e05?w=500&q=80"],
    "Chains": ["https://images.pexels.com/photos/17833829/pexels-photo-17833829.jpeg?auto=compress&w=500"],
    "Cufflinks": ["https://images.unsplash.com/photo-1726507367666-08c5f025bdf6?w=500&q=80"],
    "Earrings": ["https://images.unsplash.com/photo-1630019925419-5fc53b4a52cf?w=500&q=80"],
    "Hath Pan": ["https://images.unsplash.com/photo-1723361656145-b481be3f9e05?w=500&q=80"],
    "Maang Tikka": ["https://images.unsplash.com/photo-1723726871280-ab921c7e60c0?w=500&q=80"],
    "Mangal Sutra": ["https://images.unsplash.com/photo-1717282924526-07a7373bb142?w=500&q=80"],
    "Necklace": ["https://images.pexels.com/photos/17833830/pexels-photo-17833830.jpeg?auto=compress&w=500"],
    "Nose Pin": ["https://images.unsplash.com/photo-1587947330318-88fcd9055420?w=500&q=80"],
    "Pendant": ["https://images.unsplash.com/photo-1689775703655-6d999e38e64c?w=500&q=80"],
    "Rings": ["https://images.unsplash.com/photo-1587947330318-88fcd9055420?w=500&q=80"],
    "Tops": ["https://images.unsplash.com/photo-1723361656146-f201d215c49c?w=500&q=80"],
    "Watchbelts": ["https://images.unsplash.com/photo-1726507367666-08c5f025bdf6?w=500&q=80"],
}

# ──── AUTH HELPERS ────
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id, "email": email, "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "access"
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> dict:
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "):
        raise HTTPException(401, "Not authenticated")
    token = auth[7:]
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(401, "Invalid token")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(401, "User not found")
        user["_id"] = str(user["_id"])
        user.pop("password_hash", None)
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")

async def get_admin_user(request: Request) -> dict:
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(403, "Admin access required")
    return user

async def get_approved_user(request: Request) -> dict:
    user = await get_current_user(request)
    if user.get("role") == "admin":
        return user
    if not user.get("approved"):
        raise HTTPException(403, "Account pending approval")
    return user

# ──── MODELS ────
class RegisterRequest(BaseModel):
    name: str
    email: str
    password: str
    business_name: str
    gst_number: str
    phone: str
    state: str
    city: str
    business_address: str
    pincode: str

class LoginRequest(BaseModel):
    email: str
    password: str

class CartItemRequest(BaseModel):
    product_id: str
    category: str
    image: str
    customizations: Dict[str, str]
    notes: Optional[str] = ""

class EnquiryRequest(BaseModel):
    notes: Optional[str] = ""

class CustomisationRequest(BaseModel):
    metal_type: str
    stone_changes: str
    size_changes: str
    special_notes: str
    reference_description: Optional[str] = ""
    file_url: Optional[str] = ""
    file_name: Optional[str] = ""

class ContactRequest(BaseModel):
    name: str
    email: str
    phone: str
    message: str

class ProductCreate(BaseModel):
    product_id: str
    category: str
    images: List[str]
    rating: float = 5.0

class ProductUpdate(BaseModel):
    product_id: Optional[str] = None
    category: Optional[str] = None
    images: Optional[List[str]] = None
    rating: Optional[float] = None

class WhatsAppOrderUpdate(BaseModel):
    status: Optional[str] = None
    priority: Optional[str] = None
    assignedTo: Optional[str] = None
    admin_notes: Optional[str] = None

# ──── AUTH ENDPOINTS ────
@api_router.post("/auth/register")
async def register(req: RegisterRequest):
    email = req.email.lower().strip()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(400, "Email already registered")
    user_doc = {
    "name": req.name,
    "email": email,
    "password_hash": hash_password(req.password),
    "business_name": req.business_name,
    "gst_number": req.gst_number,
    "phone": req.phone,
    "state": req.state,
    "city": req.city,
    "business_address": req.business_address,
    "pincode": req.pincode,
    "role": "retailer",
    "approved": False,
    "created_at": datetime.now(timezone.utc).isoformat()
}
    result = await db.users.insert_one(user_doc)
    logger.info(f"New retailer registration: {req.name} ({email})")
    return {
        "message": "Registration successful! Your account is pending approval by the admin.",
        "user": {"_id": str(result.inserted_id), "name": req.name, "email": email, "role": "retailer", "approved": False}
    }

@api_router.post("/auth/login")
async def login(req: LoginRequest):
    email = req.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(req.password, user["password_hash"]):
        raise HTTPException(401, "Invalid email or password")
    if user.get("role") == "retailer" and not user.get("approved"):
        raise HTTPException(403, "Your account is pending approval. Please wait for admin verification.")
    user_id = str(user["_id"])
    token = create_access_token(user_id, email, user["role"])
    return {
        "token": token,
        "user": {
            "_id": user_id, "name": user["name"], "email": user["email"],
            "role": user["role"], "approved": user.get("approved", False),
            "business_name": user.get("business_name", ""),
            "phone": user.get("phone", ""),
        }
    }

@api_router.get("/auth/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    return {"user": user}

# ──── CATEGORIES ────
@api_router.get("/categories")
async def get_categories():
    cats = []

    for cat in CATEGORIES:

        count = await db.products.count_documents({
            "category": cat["name"]
        })

        # Check if admin has uploaded a custom Collection image
        category_image = await db.category_images.find_one(
            {"slug": cat["slug"]},
            {"_id": 0, "image": 1}
        )

        if category_image and category_image.get("image"):
            image = category_image["image"]
        else:
            # Use the existing default image if no custom image exists
            image = get_default_category_image(cat)

        cats.append({
            "name": cat["name"],
            "slug": cat["slug"],
            "prefix": cat["prefix"],
            "image": image,
            "product_count": count
        })

    return {
        "categories": cats
    }

# ──── PRODUCTS ────
@api_router.get("/products")
async def get_products(category: Optional[str] = None, page: int = 1, limit: int = 30, search: Optional[str] = None):
    query = {}
    if category:
        category_names = CATEGORY_ALIASES.get(category, [category])
        query["category"] = {"$in": category_names}
    if search:
        query["product_id"] = {"$regex": search, "$options": "i"}
    skip = (page - 1) * limit
    total = await db.products.count_documents(query)
    products = await db.products.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    return {"products": products, "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit)}

@api_router.get("/products/{product_id}")
async def get_product(product_id: str):
    product = await db.products.find_one({"product_id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(404, "Product not found")
    return {"product": product}

# ──── CART ────
@api_router.get("/cart")
async def get_cart(request: Request):
    user = await get_approved_user(request)
    cart = await db.carts.find_one({"user_id": user["_id"]}, {"_id": 0})
    return {"items": cart.get("items", []) if cart else []}

@api_router.post("/cart")
async def add_to_cart(request: Request, item: CartItemRequest):
    user = await get_approved_user(request)
    item_doc = {"item_id": str(uuid.uuid4()), **item.model_dump(), "added_at": datetime.now(timezone.utc).isoformat()}
    await db.carts.update_one({"user_id": user["_id"]}, {"$push": {"items": item_doc}}, upsert=True)
    return {"message": "Added to cart", "item": item_doc}

@api_router.delete("/cart/{item_id}")
async def remove_from_cart(request: Request, item_id: str):
    user = await get_approved_user(request)
    await db.carts.update_one({"user_id": user["_id"]}, {"$pull": {"items": {"item_id": item_id}}})
    return {"message": "Removed from cart"}

@api_router.delete("/cart")
async def clear_cart(request: Request):
    user = await get_approved_user(request)
    await db.carts.delete_one({"user_id": user["_id"]})
    return {"message": "Cart cleared"}

# ──── ENQUIRIES ────
@api_router.post("/enquiries")
async def submit_enquiry(request: Request, req: EnquiryRequest):
    user = await get_approved_user(request)
    cart = await db.carts.find_one({"user_id": user["_id"]}, {"_id": 0})
    if not cart or not cart.get("items"):
        raise HTTPException(400, "Cart is empty")
    enquiry_id = f"ENQ-{random.randint(100000, 999999)}"
    enquiry_doc = {
        "enquiry_id": enquiry_id, "user_id": user["_id"],
        "user_name": user.get("name", ""), "user_email": user.get("email", ""),
        "user_phone": user.get("phone", ""), "items": cart["items"],
        "notes": req.notes, "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.enquiries.insert_one(enquiry_doc)
    await db.carts.delete_one({"user_id": user["_id"]})
    logger.info(f"NEW ENQUIRY [{enquiry_id}] from {user.get('name')} ({user.get('email')}) - {len(cart['items'])} items")
    return {"message": "Enquiry submitted successfully! We will contact you shortly.", "enquiry_id": enquiry_id}

@api_router.get("/enquiries")
async def get_enquiries(request: Request):
    user = await get_approved_user(request)
    enquiries = await db.enquiries.find({"user_id": user["_id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return {"enquiries": enquiries}

# ──── FILE UPLOAD ────
@api_router.post("/upload")
async def upload_file(request: Request, file: UploadFile = File(...)):
    user = await get_approved_user(request)

    ext = file.filename.split(".")[-1] if "." in file.filename else "bin"

    filename = f"uploads/{user['_id']}/{uuid.uuid4()}.{ext}"

    data = await file.read()

    import io

    file_url = upload_to_s3(
        io.BytesIO(data),
        filename,
        file.content_type or "application/octet-stream",
    )

    file_doc = {
        "id": str(uuid.uuid4()),
        "storage_path": file_url,
        "original_filename": file.filename,
        "content_type": file.content_type,
        "size": len(data),
        "user_id": user["_id"],
        "is_deleted": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    await db.files.insert_one(file_doc)

    logger.info(f"File uploaded to S3: {file.filename}")

    return {
        "path": file_url,
        "filename": file.filename,
        "size": len(data),
    }

@api_router.get("/files/{path:path}")
async def download_file(path: str):
    try:
        data, content_type = get_object(path)
        return Response(content=data, media_type=content_type)
    except Exception:
        raise HTTPException(404, "File not found")

# ──── CUSTOMISATION ────
@api_router.post("/customisation")
async def submit_customisation(request: Request, req: CustomisationRequest):
    user = await get_approved_user(request)
    custom_id = f"CST-{random.randint(100000, 999999)}"
    doc = {
        "custom_id": custom_id, "user_id": user["_id"],
        "user_name": user.get("name", ""), "user_email": user.get("email", ""),
        "user_phone": user.get("phone", ""), **req.model_dump(),
        "status": "pending", "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.customisation_requests.insert_one(doc)
    logger.info(f"NEW CUSTOMISATION [{custom_id}] from {user.get('name')}")
    return {"message": "Customisation request submitted successfully!", "custom_id": custom_id}

# ──── CONTACT ────
@api_router.post("/contact")
async def submit_contact(req: ContactRequest):
    doc = {"contact_id": f"CON-{random.randint(100000, 999999)}", **req.model_dump(), "created_at": datetime.now(timezone.utc).isoformat()}
    await db.contacts.insert_one(doc)
    logger.info(f"NEW CONTACT from {req.name} ({req.email})")
    return {"message": "Message sent successfully! We will get back to you soon."}

# ──── ADMIN ────
@api_router.get("/admin/stats")
async def admin_stats(request: Request):
    await get_admin_user(request)
    return {
        "total_products": await db.products.count_documents({}),
        "total_retailers": await db.users.count_documents({"role": "retailer"}),
        "pending_approvals": await db.users.count_documents({"role": "retailer", "approved": False}),
        "total_enquiries": await db.enquiries.count_documents({}),
        "total_customisations": await db.customisation_requests.count_documents({}),
    }

@api_router.get("/admin/retailers")
async def admin_get_retailers(request: Request, status: Optional[str] = None):
    await get_admin_user(request)
    query = {"role": "retailer"}
    if status == "pending":
        query["approved"] = False
    elif status == "approved":
        query["approved"] = True
    retailers = []
    async for user in db.users.find(query, {"password_hash": 0}):
        user["_id"] = str(user["_id"])
        retailers.append(user)
    return {"retailers": retailers}

@api_router.put("/admin/retailers/{user_id}/approve")
async def approve_retailer(request: Request, user_id: str):
    await get_admin_user(request)
    result = await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": {"approved": True}})
    if result.modified_count == 0:
        raise HTTPException(404, "Retailer not found")
    logger.info(f"Retailer approved: {user_id}")
    return {"message": "Retailer approved successfully"}

@api_router.put("/admin/retailers/{user_id}/reject")
async def reject_retailer(request: Request, user_id: str):
    await get_admin_user(request)
    result = await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": {"approved": False}})
    if result.modified_count == 0:
        raise HTTPException(404, "Retailer not found")
    return {"message": "Retailer rejected"}

@api_router.get("/admin/products")
async def admin_get_products(request: Request, category: Optional[str] = None, page: int = 1, limit: int = 30):
    await get_admin_user(request)
    query = {"category": category} if category else {}
    skip = (page - 1) * limit
    total = await db.products.count_documents(query)
    products = await db.products.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    return {"products": products, "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit)}

import io

@api_router.post("/admin/products/upload")
async def admin_upload_product(
    request: Request,
    product_id: str = Form(...),
    category: str = Form(...),
    file: UploadFile = File(...)
):
    await get_admin_user(request)
    logger.info(f"product_id={product_id}")
    logger.info(f"category={category}")

    ext = file.filename.split(".")[-1] if "." in file.filename else "bin"

    filename = f"products/{uuid.uuid4()}.{ext}"

    data = await file.read()

    image_url = upload_to_s3(
    io.BytesIO(data),
    filename,
    file.content_type or "application/octet-stream",
)
    logger.info(f"S3 URL = {image_url}")
    existing = await db.products.find_one({"product_id": product_id})

    if existing:
        await db.products.update_one(
        {"product_id": product_id},
        {
            "$push": {"images": image_url},
            "$set": {"category": category}
        }
    )
    else:
        await db.products.insert_one({
        "product_id": product_id,
        "category": category,
        "category_slug": category.lower().replace(" ", "-"),
        "images": [image_url],
        "rating": 5,
        "created_at": datetime.now(timezone.utc).isoformat()
    })

    logger.info(f"Product image uploaded: {image_url}")

    return {
    "success": True,
    "message": "Product uploaded successfully",
    "url": image_url,
    "image": image_url,
    "path": image_url
}

@api_router.post("/admin/products")
async def admin_add_product(request: Request, product: ProductCreate):
    await get_admin_user(request)
    existing = await db.products.find_one({"product_id": product.product_id})
    if existing:
        raise HTTPException(400, "Product ID already exists")
    cat = next((c for c in CATEGORIES if c["name"] == product.category), None)
    if not cat:
        raise HTTPException(400, "Invalid category")
    doc = {**product.model_dump(), "category_slug": cat["slug"], "created_at": datetime.now(timezone.utc).isoformat()}
    await db.products.insert_one(doc)
    return {"message": "Product added successfully"}




@api_router.put("/admin/products/{product_id}")
async def admin_update_product(request: Request, product_id: str, update: ProductUpdate):
    await get_admin_user(request)
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(400, "No update data")
    result = await db.products.update_one({"product_id": product_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(404, "Product not found")
    return {"message": "Product updated"}


# ============================================================
# PRODUCT DETAILS / SPECIFICATIONS
# ============================================================

@api_router.put("/admin/products/{product_id}/details")
async def admin_update_product_details(
    request: Request,
    product_id: str,
    details: Dict
):
    await get_admin_user(request)

    product = await db.products.find_one(
        {"product_id": product_id}
    )

    if not product:
        raise HTTPException(
            status_code=404,
            detail="Product not found"
        )

    # Save the product-specific specifications
    await db.products.update_one(
        {"product_id": product_id},
        {
            "$set": {
                "product_details": details,
                "product_details_updated_at": datetime.now(
                    timezone.utc
                ).isoformat()
            }
        }
    )

    return {
        "success": True,
        "message": "Product details saved successfully",
        "product_id": product_id,
        "product_details": details
    }

@api_router.put("/admin/products/{product_id}/front-image")
async def admin_set_product_front_image(
    request: Request,
    product_id: str,
    image_url: str = Form(...)
):
    await get_admin_user(request)

    product = await db.products.find_one({"product_id": product_id})

    if not product:
        raise HTTPException(404, "Product not found")

    images = product.get("images", [])

    if image_url not in images:
        raise HTTPException(400, "Image does not belong to this product")

    images.remove(image_url)
    images.insert(0, image_url)

    await db.products.update_one(
        {"product_id": product_id},
        {"$set": {"images": images}}
    )

    return {
        "success": True,
        "message": "Front image updated successfully",
        "images": images
    }

@api_router.delete("/admin/products/{product_id}/image")
async def admin_delete_product_image(
    request: Request,
    product_id: str,
    image_url: str = Query(...)
):
    await get_admin_user(request)

    product = await db.products.find_one({"product_id": product_id})

    if not product:
        raise HTTPException(404, "Product not found")

    images = product.get("images", [])

    if image_url not in images:
        raise HTTPException(404, "Image not found for this product")

    # Do not allow deleting the last image
    if len(images) <= 1:
        raise HTTPException(400, "A product must have at least one image")

    images.remove(image_url)

    await db.products.update_one(
        {"product_id": product_id},
        {"$set": {"images": images}}
    )

    return {
        "success": True,
        "message": "Product image deleted successfully",
        "images": images
    }

# REPLACE PRODUCT IMAGE
@api_router.put("/admin/products/{product_id}/replace-image")
async def admin_replace_product_image(
    request: Request,
    product_id: str,
    old_image_url: str = Form(...),
    category: str = Form(...),
    file: UploadFile = File(...)
):
    await get_admin_user(request)

    product = await db.products.find_one({"product_id": product_id})

    if not product:
        raise HTTPException(404, "Product not found")

    images = product.get("images", [])

    if old_image_url not in images:
        raise HTTPException(
            404,
            "Old image not found for this product"
        )

    # Upload the new image to S3
    ext = file.filename.split(".")[-1] if "." in file.filename else "bin"

    filename = f"products/{uuid.uuid4()}.{ext}"

    data = await file.read()

    image_url = upload_to_s3(
        BytesIO(data),
        filename,
        file.content_type or "application/octet-stream"
    )

    # Find the exact position of the old image
    image_index = images.index(old_image_url)

    # Replace old image with new image
    images[image_index] = image_url

    # Save updated image list to MongoDB
    await db.products.update_one(
        {"product_id": product_id},
        {
            "$set": {
                "images": images,
                "category": category
            }
        }
    )

    logger.info(
        f"Product image replaced: "
        f"product_id={product_id}, "
        f"old_image={old_image_url}, "
        f"new_image={image_url}"
    )

    return {
        "success": True,
        "message": "Product image replaced successfully",
        "images": images,
        "image": image_url
    }

@api_router.delete("/admin/products/{product_id}")
async def admin_delete_product(request: Request, product_id: str):
    await get_admin_user(request)
    result = await db.products.delete_one({"product_id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(404, "Product not found")
    return {"message": "Product deleted"}

# ============================================================
# CATEGORY / COLLECTION IMAGES
# ============================================================

def get_category_by_slug(slug: str):
    return next(
        (cat for cat in CATEGORIES if cat["slug"] == slug),
        None
    )


def get_default_category_image(category):
    """
    Return the existing default image for a category.
    """

    aliases = CATEGORY_ALIASES.get(category["name"], [])

    # Try exact category name
    if category["name"] in STOCK_IMAGES:
        images = STOCK_IMAGES[category["name"]]
        if images:
            return images[0]

    # Try aliases such as Bangle, Earrings, Rings, etc.
    for alias in aliases:
        if alias in STOCK_IMAGES:
            images = STOCK_IMAGES[alias]
            if images:
                return images[0]

    return ""


@api_router.get("/admin/category-images")
async def admin_get_category_images(request: Request):
    await get_admin_user(request)

    category_images = {}

    async for doc in db.category_images.find({}, {"_id": 0}):
        category_images[doc["slug"]] = {
            "name": doc["name"],
            "slug": doc["slug"],
            "image": doc.get("image", ""),
            "updated_at": doc.get("updated_at")
        }

    result = []

    for category in CATEGORIES:
        existing = category_images.get(category["slug"])

        result.append({
            "name": category["name"],
            "slug": category["slug"],
            "image": (
                existing["image"]
                if existing and existing.get("image")
                else get_default_category_image(category)
            ),
            "custom_image": bool(
                existing and existing.get("image")
            ),
            "updated_at": (
                existing.get("updated_at")
                if existing
                else None
            )
        })

    return {
        "category_images": result
    }


@api_router.post("/admin/category-images/{slug}")
async def admin_upload_category_image(
    request: Request,
    slug: str,
    file: UploadFile = File(...)
):
    await get_admin_user(request)

    category = get_category_by_slug(slug)

    if not category:
        raise HTTPException(
            status_code=404,
            detail="Category not found"
        )

    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(
            status_code=400,
            detail="Only image files are allowed"
        )

    data = await file.read()

    if not data:
        raise HTTPException(
            status_code=400,
            detail="Uploaded image is empty"
        )

    if len(data) > 25 * 1024 * 1024:
        raise HTTPException(
            status_code=400,
            detail="Image must be under 25MB"
        )

    ext = (
        file.filename.split(".")[-1].lower()
        if file.filename and "." in file.filename
        else "jpg"
    )

    filename = f"category-images/{slug}/{uuid.uuid4()}.{ext}"

    image_url = upload_to_s3(
        BytesIO(data),
        filename,
        file.content_type
    )

    await db.category_images.update_one(
        {"slug": slug},
        {
            "$set": {
                "name": category["name"],
                "slug": slug,
                "image": image_url,
                "filename": filename,
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            "$setOnInsert": {
                "created_at": datetime.now(timezone.utc).isoformat()
            }
        },
        upsert=True
    )

    logger.info(
        f"Category image updated: {category['name']} -> {image_url}"
    )

    return {
        "success": True,
        "message": f"{category['name']} image updated successfully",
        "image": image_url,
        "slug": slug
    }


@api_router.delete("/admin/category-images/{slug}")
async def admin_delete_category_image(
    request: Request,
    slug: str
):
    await get_admin_user(request)

    category = get_category_by_slug(slug)

    if not category:
        raise HTTPException(
            status_code=404,
            detail="Category not found"
        )

    existing = await db.category_images.find_one({
        "slug": slug
    })

    if not existing:
        raise HTTPException(
            status_code=404,
            detail="No custom image exists for this category"
        )

    await db.category_images.delete_one({
        "slug": slug
    })

    logger.info(
        f"Category custom image deleted: {category['name']}"
    )

    return {
        "success": True,
        "message": f"{category['name']} image deleted successfully",
        "slug": slug,
        "fallback_image": get_default_category_image(category)
    }

@api_router.get("/admin/enquiries")
async def admin_get_enquiries(request: Request):
    await get_admin_user(request)
    enquiries = await db.enquiries.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return {"enquiries": enquiries}

@api_router.get("/admin/customisations")
async def admin_get_customisations(request: Request):
    await get_admin_user(request)
    customs = await db.customisation_requests.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return {"customisations": customs}

@api_router.get("/admin/whatsapp-orders")
async def admin_get_whatsapp_orders(request: Request):
    await get_admin_user(request)

    orders = await whatsapp_orders.find(
        {},
        {"_id": 0}
    ).sort("createdAt", -1).to_list(1000)

    return {
        "orders": orders
    }


# ============================================================
# WHATSAPP ORDER ANALYSIS
# ============================================================

# ============================================================
# COMBINED BUSINESS ANALYSIS
# Website + WhatsApp
# ============================================================

@api_router.get("/admin/analysis")
async def admin_combined_analysis(
    request: Request,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    channel: Optional[str] = None,
    retailer_id: Optional[str] = None,
    category: Optional[str] = None,
    product_id: Optional[str] = None,
    order_type: Optional[str] = None,
    metal: Optional[str] = None,
    purity: Optional[str] = None,
    stone: Optional[str] = None,
):
    await get_admin_user(request)

# --------------------------------------------------------
# VALIDATE DATE RANGE
# --------------------------------------------------------

    if from_date and to_date:

        if from_date > to_date:
            raise HTTPException(
                status_code=400,
                detail="From date cannot be after To date."
            )

    # --------------------------------------------------------
    # HELPERS
    # --------------------------------------------------------

    def clean(value):
        if value is None:
            return ""
        return str(value).strip()
    
    def normalize_analysis_value(value):
        value = clean(value)

        if not value:
            return ""

        return " ".join(value.split()).strip().lower()
        

    def normalize_channel(value):
        value = clean(value).lower()

        if value == "whatsapp":
            return "whatsapp"

        if value == "website":
            return "website"

        return ""

    def normalize_order_type(value):
        value = clean(value).lower()

        if value in {"custom", "customisation", "customization"}:
            return "custom"

        if value in {"stock", "catalogue", "catalog", "catalogue order"}:
            return "stock"

        return value

    def get_order_date(order):
        value = (
            order.get("order_date")
            or order.get("created_at")
            or order.get("createdAt")
        )

        if not value:
            return ""

        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d")

        return str(value)[:10]

    def get_due_date(order):
        value = order.get("due_date")

        if not value:
            return ""

        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d")

        return str(value)[:10]

    # --------------------------------------------------------
    # 1. GET WEBSITE ORDERS
    #
    # Current website submission system stores these as
    # enquiries in db.enquiries.
    # --------------------------------------------------------

    website_docs = await db.enquiries.find(
        {},
        {"_id": 0}
    ).to_list(length=None)

    website_orders = []

    for enquiry in website_docs:

        items = enquiry.get("items") or []

        normalized_items = []

        for item in items:

            normalized_items.append({
                "product_id": clean(
                    item.get("product_id")
                    or item.get("design_number")
                ),

                "design_number": clean(
                    item.get("design_number")
                    or item.get("product_id")
                ),

                "category": clean(
                    item.get("category")
                    or item.get("product_category")
                ),

                "metal": clean(item.get("metal")),

                "purity": clean(
                    item.get("purity")
                    or item.get("gold_kt")
                ),

                "gold_colour": clean(
                    item.get("gold_colour")
                    or item.get("gold_color")
                ),

                "stone": clean(
                    item.get("stone")
                    or item.get("stone_type")
                ),
            })

        website_orders.append({
            "order_id": clean(enquiry.get("enquiry_id")),
            "channel": "website",

            "retailer_id": clean(
                enquiry.get("retailer_id")
                or enquiry.get("user_id")
            ),

            "retailer_name": clean(
                enquiry.get("retailer_name")
                or enquiry.get("user_name")
            ),

            "order_date": get_order_date(enquiry),
            "due_date": get_due_date(enquiry),

            "order_type": normalize_order_type(
                enquiry.get("order_type")
            ),

            "status": clean(
                enquiry.get("status")
            ),

            "items": normalized_items,
        })

    # --------------------------------------------------------
    # 2. GET WHATSAPP ORDERS
    # --------------------------------------------------------

    whatsapp_docs = await whatsapp_orders.find(
        {},
        {"_id": 0}
    ).to_list(length=None)

    whatsapp_normalized = []

    for order in whatsapp_docs:

        item = {
            "product_id": clean(
                order.get("product_id")
                or order.get("design_number")
            ),

            "design_number": clean(
                order.get("design_number")
                or order.get("product_id")
            ),

            "category": clean(
                order.get("product_category")
                or order.get("category")
            ),

            "metal": clean(
                order.get("metal")
            ),

            "purity": clean(
                order.get("gold_kt")
                or order.get("purity")
            ),

            "gold_colour": clean(
                order.get("gold_colour")
                or order.get("gold_color")
            ),

            "stone": clean(
                order.get("stone_type")
                or order.get("stone")
            ),
        }

        whatsapp_normalized.append({
            "order_id": clean(
                order.get("orderId")
                or order.get("order_id")
            ),

            "channel": "whatsapp",

            "retailer_id": clean(
                order.get("retailer_id")
            ),

            "retailer_name": clean(
                order.get("retailer_name")
            ),

            "order_date": get_order_date(order),
            "due_date": get_due_date(order),

            "order_type": normalize_order_type(
                order.get("order_type")
            ),

            "status": clean(
                order.get("status")
            ),

            "items": [item],
        })

    # --------------------------------------------------------
    # 3. COMBINE WEBSITE + WHATSAPP
    # --------------------------------------------------------

    all_orders = website_orders + whatsapp_normalized

    # --------------------------------------------------------
    # 4. FILTER ORDERS
    # --------------------------------------------------------

    filtered_orders = []

    for order in all_orders:

        order_date = clean(order.get("order_date"))

        # Date
        if from_date or to_date:

            # If a date filter is selected, orders without
            # a valid order date must not be included.
            if not order_date:
                continue

            if from_date and order_date < from_date:
                continue

            if to_date and order_date > to_date:
                continue

        # Channel
        if channel and channel.lower() != "all":
            if clean(order.get("channel")).lower() != channel.lower():
                continue

        # Retailer
        if retailer_id and retailer_id.lower() != "all":
            if clean(order.get("retailer_id")) != retailer_id:
                continue

        # Order type
        if order_type and order_type.lower() != "all":

            requested_type = normalize_order_type(order_type)

            if normalize_order_type(order.get("order_type")) != requested_type:
                continue

        # Item-level filters
        matching_items = []

        for item in order.get("items", []):

            if category and category.lower() != "all":

                if normalize_analysis_value(
                    item.get("category")
                ) != normalize_analysis_value(category):
                    continue

            if product_id and product_id.lower() != "all":

                requested_product = normalize_analysis_value(
                    product_id
                )

                item_product_id = normalize_analysis_value(
                    item.get("product_id")
                )

                item_design_number = normalize_analysis_value(
                    item.get("design_number")
                )

                if (
                    item_product_id != requested_product
                    and item_design_number != requested_product
                ):
                    continue

            if metal and metal.lower() != "all":

                requested_metal = normalize_analysis_value(
                    metal
                )

                item_metal = normalize_analysis_value(
                    item.get("metal")
                )

                if "gold" in requested_metal and "platinum" in requested_metal:

                    if not (
                        "gold" in item_metal
                        and "platinum" in item_metal
                    ):
                        continue

                elif requested_metal == "gold":

                    if (
                        "gold" not in item_metal
                        or "platinum" in item_metal
                    ):
                        continue

                elif requested_metal == "platinum":

                    if "platinum" not in item_metal:
                        continue

                elif item_metal != requested_metal:
                    continue

            if purity and purity.lower() != "all":

                if normalize_analysis_value(
                    item.get("purity")
                ) != normalize_analysis_value(purity):
                        continue

            if stone and stone.lower() != "all":

                if normalize_analysis_value(
                    item.get("stone")
                ) != normalize_analysis_value(stone):
                        continue

            matching_items.append(item)

        # If item filters were supplied, order must contain
        # at least one matching item.
        item_filter_used = any([
            category and category.lower() != "all",
            product_id and product_id.lower() != "all",
            metal and metal.lower() != "all",
            purity and purity.lower() != "all",
            stone and stone.lower() != "all",
        ])

        if item_filter_used:
            if not matching_items:
                continue

            order["items"] = matching_items

        filtered_orders.append(order)

    orders = filtered_orders

    # --------------------------------------------------------
    # 5. OVERVIEW
    # --------------------------------------------------------

    total_orders = len(orders)
    combined_orders = total_orders

    website_order_count = sum(
        1
        for order in orders
        if order.get("channel") == "website"
    )

    whatsapp_order_count = sum(
        1
        for order in orders
        if order.get("channel") == "whatsapp"
    )

    total_products = sum(
        len(order.get("items", []))
        for order in orders
    )

    dates = {
        order.get("order_date")
        for order in orders
        if order.get("order_date")
    }

    if from_date and to_date:
        try:
            start = datetime.strptime(from_date, "%Y-%m-%d")
            end = datetime.strptime(to_date, "%Y-%m-%d")
            number_of_days = max(1, (end - start).days + 1)
        except Exception:
            number_of_days = max(1, len(dates))
    else:
        number_of_days = max(1, len(dates))

    average_orders_per_day = (
        round(total_orders / number_of_days, 2)
        if number_of_days
        else 0
    )

    # --------------------------------------------------------
    # 6. HELPER FOR ITEM GROUPING
    # --------------------------------------------------------

    def group_items(field):
        result = {}

        for order in orders:
            for item in order.get("items", []):

                value = clean(item.get(field))

                if not value:
                    continue

                result[value] = result.get(value, 0) + 1

        return [
            {
                "name": name,
                "count": count
            }
            for name, count in sorted(
                result.items(),
                key=lambda x: x[1],
                reverse=True
            )
        ]

    # --------------------------------------------------------
    # 7. CATEGORY
    # --------------------------------------------------------

    category_data = group_items("category")

    # --------------------------------------------------------
    # 8. PRODUCT / DESIGN
    # --------------------------------------------------------

    product_counts = {}

    for order in orders:
        for item in order.get("items", []):

            product_id = clean(
                item.get("product_id")
                or item.get("design_number")
            )

            design_number = clean(
                item.get("design_number")
                or item.get("product_id")
            )

            product_key = normalize_analysis_value(
    product_id or design_number
)

            if not product_key:
                continue

            if product_key not in product_counts:
                product_counts[product_key] = {
                    "product_id": product_id,
                    "design_number": design_number,
                    "category": clean(item.get("category")),
                    "orders": 0,
                }

            product_counts[product_key]["orders"] += 1


    # Get the COMPLETE catalogue
    catalogue_products = await db.products.find(
        {},
        {"_id": 0}
    ).to_list(length=None)


    product_performance = []

    for product in catalogue_products:

        product_id = clean(
            product.get("product_id")
            or product.get("design_number")
        )

        design_number = clean(
            product.get("design_number")
            or product.get("product_id")
        )

        product_key = product_id or design_number

        if not product_key:
            continue

        ordered = product_counts.get(product_key)

        product_performance.append({
            "product_id": product_id,
            "design_number": design_number,
            "name": clean(product.get("name")),
            "category": clean(product.get("category")),
            "orders": ordered["orders"] if ordered else 0,
        })


    # Best sellers
    best_sellers = sorted(
        product_performance,
        key=lambda x: x["orders"],
        reverse=True
    )


        # Underperforming = products that have orders,
    # but are among the lowest-selling products.
    ordered_products = [
        product
        for product in product_performance
        if product["orders"] > 0
    ]

    ordered_products.sort(
        key=lambda x: x["orders"]
    )

    underperforming_products = ordered_products[:20]


    # Never ordered
    never_ordered_products = [
        product
        for product in product_performance
        if product["orders"] == 0
    ]


    # Keep the complete product list available
    product_data = sorted(
        product_performance,
        key=lambda x: x["orders"],
        reverse=True
    )

        # --------------------------------------------------------
    # METAL ANALYSIS
    # --------------------------------------------------------

    metal_counts = {}

    for order in orders:
        for item in order.get("items", []):

            raw_metal = normalize_analysis_value(
                item.get("metal")
            )

            if not raw_metal:
                continue

            # Normalize all common variations
            if "gold" in raw_metal and "platinum" in raw_metal:
                metal_name = "Gold + Platinum"

            elif "platinum" in raw_metal:
                metal_name = "Platinum"

            elif "gold" in raw_metal:
                metal_name = "Gold"

            else:
                # Keep unexpected values rather than losing data
                metal_name = clean(item.get("metal"))

            metal_counts[metal_name] = (
                metal_counts.get(metal_name, 0) + 1
            )

    metal_data = [
        {
            "name": metal_name,
            "orders": order_count
        }
        for metal_name, order_count
        in metal_counts.items()
    ]

    # Keep the requested metal groups in a consistent order
    metal_order = {
        "Gold": 1,
        "Platinum": 2,
        "Gold + Platinum": 3
    }

    metal_data.sort(
        key=lambda x: (
            metal_order.get(x["name"], 99),
            x["name"]
        )
    )

    # --------------------------------------------------------
    # 10. PURITY
    # --------------------------------------------------------

    purity_data = group_items("purity")

    # --------------------------------------------------------
    # 11. GOLD COLOUR
    # --------------------------------------------------------

    gold_colour_data = group_items("gold_colour")

    # --------------------------------------------------------
    # STONE ANALYSIS
    # --------------------------------------------------------

    stone_counts = {}

    for order in orders:
        for item in order.get("items", []):

            raw_stone = normalize_analysis_value(
                item.get("stone")
            )

            if not raw_stone:
                continue

            # Normalize common stone variations
            if (
                "natural" in raw_stone
                and "diamond" in raw_stone
            ):
                stone_name = "Natural Diamond"

            elif (
                "lab" in raw_stone
                and "diamond" in raw_stone
            ):
                stone_name = "Lab Grown"

            elif raw_stone in {
                "cz",
                "c.z.",
                "cubic zirconia",
                "cubic zircon",
            }:
                stone_name = "CZ"

            elif (
                "colour stone" in raw_stone
                or "color stone" in raw_stone
                or "coloured stone" in raw_stone
                or "colored stone" in raw_stone
            ):
                stone_name = "Colour Stones"

            elif "precious stone" in raw_stone:
                stone_name = "Precious Stones"

            else:
                # Preserve unknown values
                stone_name = clean(item.get("stone"))

            stone_counts[stone_name] = (
                stone_counts.get(stone_name, 0) + 1
            )

    stone_data = [
        {
            "name": stone_name,
            "orders": order_count
        }
        for stone_name, order_count
        in stone_counts.items()
    ]

    stone_data.sort(
        key=lambda x: x["orders"],
        reverse=True
    )


        # --------------------------------------------------------
    # 12A. CATEGORY -> PRODUCT / DESIGN DRILL-DOWN
    # --------------------------------------------------------

    category_product_map = {}

    # First build the complete catalogue structure
    # so products with zero orders are also available.

    for product in products:

        category_name = normalize_analysis_value(
            product.get("category")
        )

        if not category_name:
            continue

        product_id = clean(
            product.get("product_id")
            or product.get("id")
            or product.get("design_number")
        )

        design_number = clean(
            product.get("design_number")
            or product.get("product_id")
            or product.get("id")
        )

        product_key = product_id or design_number

        if not product_key:
            continue

        category_product_map.setdefault(
            category_name,
            {}
        )

        category_product_map[category_name].setdefault(
            product_key,
            {
                "product_id": product_id,
                "design_number": design_number,
                "orders": 0
            }
        )


    # Now add actual orders

    for order in orders:

        for item in order.get("items", []):

            category_name = normalize_analysis_value(
                item.get("category")
            )

            if not category_name:
                continue

            product_id = clean(
                item.get("product_id")
                or item.get("design_number")
            )

            design_number = clean(
                item.get("design_number")
                or item.get("product_id")
            )

            product_key = product_id or design_number

            if not product_key:
                continue

            category_product_map.setdefault(
                category_name,
                {}
            )

            category_product_map[category_name].setdefault(
                product_key,
                {
                    "product_id": product_id,
                    "design_number": design_number,
                    "orders": 0
                }
            )

            category_product_map[category_name][
                product_key
            ]["orders"] += 1


    # Convert dictionaries into frontend-friendly arrays

    category_product_data = {}

    for category_name, products_map in category_product_map.items():

        category_product_data[category_name] = sorted(
            products_map.values(),
            key=lambda x: (
                -x["orders"],
                x["product_id"] or x["design_number"]
            )
        )


    # --------------------------------------------------------
    # 12B. CATEGORY × METAL
    # --------------------------------------------------------

    category_metal_map = {}

    for order in orders:
        for item in order.get("items", []):

            category_name = normalize_analysis_value(
    item.get("category")
)

            metal_name = clean(
                item.get("metal")
            )

            if not category_name or not metal_name:
                continue

            category_metal_map.setdefault(
                category_name,
                {}
            )

            category_metal_map[category_name].setdefault(
                metal_name,
                {
                    "orders": 0,
                    "products": {}
                }
            )

            metal_data_item = category_metal_map[
                category_name
            ][metal_name]

            metal_data_item["orders"] += 1

            product_id = clean(
                item.get("product_id")
                or item.get("design_number")
            )

            design_number = clean(
                item.get("design_number")
                or item.get("product_id")
            )

            product_key = product_id or design_number

            if product_key:

                if product_key not in metal_data_item["products"]:
                    metal_data_item["products"][product_key] = {
                        "product_id": product_id,
                        "design_number": design_number,
                        "orders": 0
                    }

                metal_data_item["products"][product_key]["orders"] += 1


    category_metal_data = {}

    for category_name, metals in category_metal_map.items():

        category_metal_data[category_name] = {}

        for metal_name, metal_info in metals.items():

            category_metal_data[category_name][metal_name] = {
                "orders": metal_info["orders"],
                "products": sorted(
                    metal_info["products"].values(),
                    key=lambda x: x["orders"],
                    reverse=True
                )
            }


    # --------------------------------------------------------
    # 12C. CATEGORY × STONE
    # --------------------------------------------------------

    category_stone_map = {}

    for order in orders:
        for item in order.get("items", []):

            category_name = normalize_analysis_value(
    item.get("category")
)

            stone_name = clean(
                item.get("stone")
            )

            if not category_name or not stone_name:
                continue

            category_stone_map.setdefault(
                category_name,
                {}
            )

            category_stone_map[category_name].setdefault(
                stone_name,
                {
                    "orders": 0,
                    "products": {}
                }
            )

            stone_data_item = category_stone_map[
                category_name
            ][stone_name]

            stone_data_item["orders"] += 1

            product_id = clean(
                item.get("product_id")
                or item.get("design_number")
            )

            design_number = clean(
                item.get("design_number")
                or item.get("product_id")
            )

            product_key = normalize_analysis_value(
    product_id or design_number
)

            if product_key:

                if product_key not in stone_data_item["products"]:
                    stone_data_item["products"][product_key] = {
                        "product_id": product_id,
                        "design_number": design_number,
                        "orders": 0
                    }

                stone_data_item["products"][product_key]["orders"] += 1


    category_stone_data = {}

    for category_name, stones in category_stone_map.items():

        category_stone_data[category_name] = {}

        for stone_name, stone_info in stones.items():

            category_stone_data[category_name][stone_name] = {
                "orders": stone_info["orders"],
                "products": sorted(
                    stone_info["products"].values(),
                    key=lambda x: x["orders"],
                    reverse=True
                )
            }


    # --------------------------------------------------------
    # DAILY ORDER ANALYSIS
    # --------------------------------------------------------

    by_date_counts = {}

    for order in orders:
        order_date = clean(order.get("order_date"))

        if not order_date:
            continue

        by_date_counts[order_date] = (
            by_date_counts.get(order_date, 0) + 1
        )

    by_date = [
        {
            "date": date,
            "count": count
        }
        for date, count in sorted(
            by_date_counts.items(),
            key=lambda x: x[0]
        )
    ]


    # --------------------------------------------------------
    # 13. STATUS
    # --------------------------------------------------------

    status_counts = {}

    for order in orders:

        value = clean(order.get("status"))

        if not value:
            continue

        status_counts[value] = (
            status_counts.get(value, 0) + 1
        )

    status_data = [
        {
            "name": name,
            "count": count
        }
        for name, count in sorted(
            status_counts.items(),
            key=lambda x: x[1],
            reverse=True
        )
    ]

    # --------------------------------------------------------
    # 14. CATEGORY ORDER %
    # --------------------------------------------------------

    for category_item in category_data:

        category_item["percentage"] = (
            round(
                (
                    category_item["count"]
                    / total_orders
                ) * 100,
                2
            )
            if total_orders
            else 0
        )

    # Combined Website + WhatsApp orders
    combined_orders = total_orders

        # --------------------------------------------------------
    # 15. MONTHLY CATEGORY PERFORMANCE + GROWTH
    # --------------------------------------------------------

    monthly_category = {}

    for order in orders:

        order_date = clean(order.get("order_date"))

        if not order_date:
            continue

        month = order_date[:7]

        for item in order.get("items", []):

            category_name = normalize_analysis_value(
    item.get("category")
)

            if not category_name:
                continue

            monthly_category.setdefault(month, {})

            monthly_category[month][category_name] = (
                monthly_category[month].get(category_name, 0) + 1
            )


    # --------------------------------------------------------
    # ALL CATEGORIES
    # --------------------------------------------------------

    all_categories = set()

    for month_data in monthly_category.values():
        all_categories.update(month_data.keys())


    # --------------------------------------------------------
    # MONTHLY DATA + GROWTH %
    # --------------------------------------------------------

    category_monthly = []

    previous_month_data = {}

        # Build a continuous month range so months with zero orders
    # are also included in the category growth chart.

    months_sorted = sorted(monthly_category)

    if months_sorted:

        first_month = datetime.strptime(
            months_sorted[0],
            "%Y-%m"
        )

        last_month = datetime.strptime(
            months_sorted[-1],
            "%Y-%m"
        )

        all_months = []

        current_month = first_month

        while current_month <= last_month:

            all_months.append(
                current_month.strftime("%Y-%m")
            )

            if current_month.month == 12:

                current_month = current_month.replace(
                    year=current_month.year + 1,
                    month=1
                )

            else:

                current_month = current_month.replace(
                    month=current_month.month + 1
                )

    else:

        all_months = []


    for month in all_months:

        current_month_data = monthly_category.get(
    month,
    {}
)

        month_categories = {}

        for category_name in sorted(all_categories):

            current_orders = current_month_data.get(
                category_name,
                0
            )

            previous_orders = previous_month_data.get(
                category_name,
                0
            )

            # ---------------------------------------------
            # GROWTH / DECLINE
            # ---------------------------------------------

            if previous_orders == 0 and current_orders > 0:

                growth_percentage = None
                growth_status = "new"

            elif previous_orders == 0 and current_orders == 0:

                growth_percentage = None
                growth_status = "no_orders"

            else:

                growth_percentage = round(
                    (
                        (current_orders - previous_orders)
                        / previous_orders
                    ) * 100,
                    2
                )

                if growth_percentage > 0:
                    growth_status = "growth"

                elif growth_percentage < 0:
                    growth_status = "decline"

                else:
                    growth_status = "unchanged"

            month_categories[category_name] = {
                "orders": current_orders,
                "growth_percentage": growth_percentage,
                "growth_status": growth_status
            }

        category_monthly.append({
            "month": month,
            "categories": month_categories
        })

        previous_month_data = current_month_data.copy()

        # --------------------------------------------------------
    # 16. RETAILER ANALYSIS
    # --------------------------------------------------------

    retailer_map = {}

    for order in orders:

        retailer_key = (
            clean(order.get("retailer_id"))
            or clean(order.get("retailer_name"))
            or "Unknown"
        )

        retailer_name = (
            clean(order.get("retailer_name"))
            or "Unknown"
        )

        if retailer_key not in retailer_map:

            retailer_map[retailer_key] = {
                "retailer_id": retailer_key,
                "retailer_name": retailer_name,

                "total_orders": 0,
                "custom_orders": 0,
                "stock_orders": 0,

                # Existing category summary
                "categories": {},

                # New category -> product/design drill-down
                "category_details": {}
            }

        retailer = retailer_map[retailer_key]

        # ---------------------------------------------
        # TOTAL / CUSTOM / STOCK ORDERS
        # ---------------------------------------------

        retailer["total_orders"] += 1

        order_type_value = normalize_order_type(
            order.get("order_type")
        )

        if order_type_value == "custom":
            retailer["custom_orders"] += 1

        elif order_type_value == "stock":
            retailer["stock_orders"] += 1

        # ---------------------------------------------
        # CATEGORY + PRODUCT / DESIGN
        # ---------------------------------------------

        for item in order.get("items", []):

            category_name = normalize_analysis_value(
    item.get("category")
)

            if not category_name:
                continue

            # Existing category count
            retailer["categories"][category_name] = (
                retailer["categories"].get(
                    category_name,
                    0
                ) + 1
            )

            # Create category detail
            if category_name not in retailer["category_details"]:

                retailer["category_details"][category_name] = {
                    "orders": 0,
                    "products": {}
                }

            category_detail = retailer["category_details"][
                category_name
            ]

            category_detail["orders"] += 1

            # Product / Design number
            product_id = clean(
                item.get("product_id")
            )

            design_number = clean(
                item.get("design_number")
            )

            product_key = normalize_analysis_value(
                design_number
                or product_id
            )

            if not product_key:
                product_key = "unknown"

            # Create product/design entry
            if product_key not in category_detail["products"]:

                category_detail["products"][product_key] = {
                    "product_id": product_id,
                    "design_number": design_number,
                    "name": clean(
                        item.get("name")
                    ),
                    "orders": 0
                }

            category_detail["products"][product_key]["orders"] += 1

    # ---------------------------------------------
    # CONVERT CATEGORY DETAILS TO FRONTEND-FRIENDLY
    # ARRAYS
    # ---------------------------------------------

    retailer_data = []

    for retailer in retailer_map.values():

        category_details = {}

        for category_name, category_info in retailer[
            "category_details"
        ].items():

            products = list(
                category_info["products"].values()
            )

            products.sort(
                key=lambda x: x["orders"],
                reverse=True
            )

            category_details[category_name] = {
                "orders": category_info["orders"],
                "products": products
            }

        retailer["category_details"] = category_details

        retailer_data.append(retailer)

    # ---------------------------------------------
    # SORT RETAILERS BY TOTAL ORDERS
    # ---------------------------------------------

    retailer_data.sort(
        key=lambda x: x["total_orders"],
        reverse=True
    )

        # --------------------------------------------------------
    # 16A. CATEGORY × METAL
    # --------------------------------------------------------

    category_metal_map = {}

    for order in orders:

        for item in order.get("items", []):

            category_name = normalize_analysis_value(
                item.get("category")
            )

            metal_name = normalize_analysis_value(
                item.get("metal")
            )

            if not category_name or not metal_name:
                continue

            # Keep the same metal grouping used in Metal Analysis
            if "gold" in metal_name and "platinum" in metal_name:
                metal_name = "Gold + Platinum"

            elif "platinum" in metal_name:
                metal_name = "Platinum"

            elif "gold" in metal_name:
                metal_name = "Gold"

            else:
                metal_name = clean(item.get("metal"))

            category_metal_map.setdefault(
                category_name,
                {}
            )

            category_metal_map[category_name].setdefault(
                metal_name,
                {
                    "orders": 0,
                    "products": {}
                }
            )

            metal_detail = category_metal_map[
                category_name
            ][metal_name]

            metal_detail["orders"] += 1

            product_id = clean(
                item.get("product_id")
                or item.get("design_number")
            )

            design_number = clean(
                item.get("design_number")
                or item.get("product_id")
            )

            product_key = normalize_analysis_value(
                design_number
                or product_id
            )

            if not product_key:
                product_key = "unknown"

            if product_key not in metal_detail["products"]:

                metal_detail["products"][product_key] = {
                    "product_id": product_id,
                    "design_number": design_number,
                    "name": clean(item.get("name")),
                    "orders": 0
                }

            metal_detail["products"][
                product_key
            ]["orders"] += 1


    # Convert to frontend-friendly arrays

    category_metal_data = {}

    for category_name, metals in category_metal_map.items():

        category_metal_data[category_name] = {}

        for metal_name, metal_info in metals.items():

            products = list(
                metal_info["products"].values()
            )

            products.sort(
                key=lambda x: x["orders"],
                reverse=True
            )

            category_metal_data[
                category_name
            ][metal_name] = {
                "orders": metal_info["orders"],
                "products": products
            }


    # --------------------------------------------------------
    # 16B. CATEGORY × STONE
    # --------------------------------------------------------

    category_stone_map = {}

    for order in orders:

        for item in order.get("items", []):

            category_name = normalize_analysis_value(
                item.get("category")
            )

            stone_name = normalize_analysis_value(
                item.get("stone")
            )

            if not category_name or not stone_name:
                continue

            # Normalize common stone names
            if (
                "natural" in stone_name
                and "diamond" in stone_name
            ):
                stone_name = "Natural Diamond"

            elif (
                "lab" in stone_name
                and "diamond" in stone_name
            ):
                stone_name = "Lab Grown"

            elif stone_name in {
                "cz",
                "c.z.",
                "cubic zirconia",
                "cubic zircon"
            }:
                stone_name = "CZ"

            elif (
                "colour stone" in stone_name
                or "color stone" in stone_name
                or "coloured stone" in stone_name
                or "colored stone" in stone_name
            ):
                stone_name = "Colour Stones"

            elif "precious stone" in stone_name:
                stone_name = "Precious Stones"

            else:
                stone_name = clean(item.get("stone"))

            category_stone_map.setdefault(
                category_name,
                {}
            )

            category_stone_map[category_name].setdefault(
                stone_name,
                {
                    "orders": 0,
                    "products": {}
                }
            )

            stone_detail = category_stone_map[
                category_name
            ][stone_name]

            stone_detail["orders"] += 1

            product_id = clean(
                item.get("product_id")
                or item.get("design_number")
            )

            design_number = clean(
                item.get("design_number")
                or item.get("product_id")
            )

            product_key = normalize_analysis_value(
                design_number
                or product_id
            )

            if not product_key:
                product_key = "unknown"

            if product_key not in stone_detail["products"]:

                stone_detail["products"][product_key] = {
                    "product_id": product_id,
                    "design_number": design_number,
                    "name": clean(item.get("name")),
                    "orders": 0
                }

            stone_detail["products"][
                product_key
            ]["orders"] += 1


    # Convert to frontend-friendly arrays

    category_stone_data = {}

    for category_name, stones in category_stone_map.items():

        category_stone_data[category_name] = {}

        for stone_name, stone_info in stones.items():

            products = list(
                stone_info["products"].values()
            )

            products.sort(
                key=lambda x: x["orders"],
                reverse=True
            )

            category_stone_data[
                category_name
            ][stone_name] = {
                "orders": stone_info["orders"],
                "products": products
            }

    # --------------------------------------------------------
    # 17. DUE DATE ANALYSIS
    # --------------------------------------------------------

    today = datetime.now(timezone.utc).date()

    due_this_week = 0
    due_next_week = 0
    overdue = 0
    completed_on_time = 0
    delayed = 0

    completed_statuses = {
    "delivered",
    "completed"
}

    for order in orders:

        due_date_string = clean(
            order.get("due_date")
        )

        if not due_date_string:
            continue

        try:
            due_date_value = datetime.strptime(
                due_date_string[:10],
                "%Y-%m-%d"
            ).date()
        except Exception:
            continue

        status = clean(
            order.get("status")
        ).lower()

        days_difference = (
            due_date_value - today
        ).days

        if status in completed_statuses:

            completed_at = order.get("completedAt")

            if completed_at:

                try:
                    if hasattr(completed_at, "date"):
                        completed_date = completed_at.date()
                    else:
                        completed_date = datetime.strptime(
                            str(completed_at)[:10],
                            "%Y-%m-%d"
                        ).date()

                    if completed_date <= due_date_value:
                        completed_on_time += 1
                    else:
                        delayed += 1

                except Exception:
                    pass

        else:

            if days_difference < 0:
                overdue += 1

            elif 0 <= days_difference <= 6:
                due_this_week += 1

            elif 7 <= days_difference <= 13:
                due_next_week += 1

    # --------------------------------------------------------
    # 18. RETURN
    # --------------------------------------------------------

        # --------------------------------------------------------
    # 18. AUTOMATIC BUSINESS INSIGHTS
    # --------------------------------------------------------

    automatic_insights = []


    # 1. Top category
    if category_data:

        top_category = max(
            category_data,
            key=lambda x: x.get("count", 0)
        )

        automatic_insights.append({
            "type": "top_category",
            "title": "Top Category",
            "message": (
                f"{top_category.get('name')} is your "
                f"most ordered category with "
                f"{top_category.get('count', 0)} orders."
            ),
            "data": {
                "category": top_category.get("name"),
                "orders": top_category.get("count", 0)
            }
        })


    # 2. Best-selling product
    if best_sellers:

        top_product = best_sellers[0]

        if top_product.get("orders", 0) > 0:

            product_name = (
                top_product.get("design_number")
                or top_product.get("product_id")
                or top_product.get("name")
                or "Unknown Product"
            )

            automatic_insights.append({
                "type": "best_seller",
                "title": "Best-Selling Design",
                "message": (
                    f"{product_name} is the top-selling design "
                    f"with {top_product.get('orders', 0)} orders."
                ),
                "data": {
                    "product_id": top_product.get("product_id"),
                    "design_number": top_product.get("design_number"),
                    "orders": top_product.get("orders", 0)
                }
            })


    # 3. Never ordered products
    never_ordered_count = len(
        never_ordered_products
    )

    if never_ordered_count > 0:

        automatic_insights.append({
            "type": "never_ordered",
            "title": "Catalogue Opportunity",
            "message": (
                f"{never_ordered_count} catalogue "
                f"designs have never been ordered."
            ),
            "data": {
                "count": never_ordered_count
            }
        })


    # 4. Top retailer
    if retailer_data:

        top_retailer = retailer_data[0]

        automatic_insights.append({
            "type": "top_retailer",
            "title": "Top Retailer",
            "message": (
                f"{top_retailer.get('retailer_name', 'Unknown')} "
                f"has placed the most orders with "
                f"{top_retailer.get('total_orders', 0)} orders."
            ),
            "data": {
                "retailer_id": top_retailer.get("retailer_id"),
                "retailer_name": top_retailer.get("retailer_name"),
                "orders": top_retailer.get("total_orders", 0)
            }
        })


    # 5. Overdue orders
    if overdue > 0:

        automatic_insights.append({
            "type": "overdue",
            "title": "Overdue Orders",
            "message": (
                f"{overdue} orders are currently overdue."
            ),
            "data": {
                "count": overdue
            }
        })


    # 6. Latest category growth / decline
    if category_monthly:

        latest_month = category_monthly[-1]

        growth_items = []

        for category_name, category_info in (
            latest_month.get("categories", {}).items()
        ):

            growth = category_info.get(
                "growth_percentage"
            )

            if growth is not None:
                growth_items.append({
                    "category": category_name,
                    "growth_percentage": growth,
                    "orders": category_info.get("orders", 0)
                })

        if growth_items:

            highest_growth = max(
                growth_items,
                key=lambda x: x["growth_percentage"]
            )

            lowest_growth = min(
                growth_items,
                key=lambda x: x["growth_percentage"]
            )

            if highest_growth["growth_percentage"] > 0:

                automatic_insights.append({
                    "type": "category_growth",
                    "title": "Fastest Growing Category",
                    "message": (
                        f"{highest_growth['category']} grew by "
                        f"{highest_growth['growth_percentage']}% "
                        f"in {latest_month.get('month')}."
                    ),
                    "data": highest_growth
                })

            if lowest_growth["growth_percentage"] < 0:

                automatic_insights.append({
                    "type": "category_decline",
                    "title": "Category Decline",
                    "message": (
                        f"{lowest_growth['category']} declined by "
                        f"{abs(lowest_growth['growth_percentage'])}% "
                        f"in {latest_month.get('month')}."
                    ),
                    "data": lowest_growth
                })

    return {
        "success": True,

        "filters": {
            "from_date": from_date,
            "to_date": to_date,
            "channel": channel or "all",
            "retailer_id": retailer_id or "all",
            "category": category or "all",
            "product_id": product_id or "all",
            "order_type": order_type or "all",
            "metal": metal or "all",
            "purity": purity or "all",
            "stone": stone or "all",
        },

        "overview": {
    "total_orders": total_orders,
    "combined_orders": combined_orders,
    "website_orders": website_order_count,
    "whatsapp_orders": whatsapp_order_count,
    "total_products": total_products,
    "average_orders_per_day": average_orders_per_day,
},

"by_date": by_date,

"category": category_data,

"category_product_drilldown": category_product_data,

"category_monthly": category_monthly,

"products": product_data,

        "product_intelligence": {
        "best_sellers": best_sellers[:20],
        "underperforming": underperforming_products[:20],
        "never_ordered": never_ordered_products,
        },

        "retailers": retailer_data,

        "metal": metal_data,

        "purity": purity_data,

        "gold_colour": gold_colour_data,

        "stone": stone_data,

"cross_analysis": {
    "category_metal": category_metal_data,
    "category_stone": category_stone_data
},

"status": status_data,

        "due_dates": {
            "due_this_week": due_this_week,
            "due_next_week": due_next_week,
            "overdue": overdue,
            "completed_on_time": completed_on_time,
            "delayed": delayed,
        },
        "automatic_insights": automatic_insights,
    }

# ============================================================
# ANALYSIS EXPORT — CSV
# ============================================================

@api_router.get("/admin/analysis/export/csv")
async def export_analysis_csv(
    request: Request,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    channel: Optional[str] = None,
    retailer_id: Optional[str] = None,
    category: Optional[str] = None,
    product_id: Optional[str] = None,
    order_type: Optional[str] = None,
    metal: Optional[str] = None,
    purity: Optional[str] = None,
    stone: Optional[str] = None,
):
    # Get the exact same filtered Analysis data
    analysis = await admin_combined_analysis(
        request=request,
        from_date=from_date,
        to_date=to_date,
        channel=channel,
        retailer_id=retailer_id,
        category=category,
        product_id=product_id,
        order_type=order_type,
        metal=metal,
        purity=purity,
        stone=stone,
    )

    output = BytesIO()

    # UTF-8 BOM so Excel opens the CSV correctly
    output.write("\ufeff".encode("utf-8"))

    # Convert bytes buffer into text writer
    import io

    text_output = io.TextIOWrapper(
        output,
        encoding="utf-8",
        newline=""
    )

    writer = csv.writer(text_output)

    # --------------------------------------------------------
    # FILTERS
    # --------------------------------------------------------

    writer.writerow(["ANALYSIS REPORT"])
    writer.writerow([])

    writer.writerow(["FILTERS"])
    writer.writerow(["From Date", analysis["filters"]["from_date"]])
    writer.writerow(["To Date", analysis["filters"]["to_date"]])
    writer.writerow(["Channel", analysis["filters"]["channel"]])
    writer.writerow(["Retailer", analysis["filters"]["retailer_id"]])
    writer.writerow(["Category", analysis["filters"]["category"]])
    writer.writerow(["Product / Design", analysis["filters"]["product_id"]])
    writer.writerow(["Custom / Stock", analysis["filters"]["order_type"]])
    writer.writerow(["Metal", analysis["filters"]["metal"]])
    writer.writerow(["Purity", analysis["filters"]["purity"]])
    writer.writerow(["Stone", analysis["filters"]["stone"]])

    writer.writerow([])

    # --------------------------------------------------------
    # OVERVIEW
    # --------------------------------------------------------

    writer.writerow(["OVERVIEW"])

    overview = analysis["overview"]

    writer.writerow([
        "Metric",
        "Value"
    ])

    writer.writerow([
        "Total Orders",
        overview["total_orders"]
    ])

    writer.writerow([
        "Combined Orders",
        overview["combined_orders"]
    ])

    writer.writerow([
        "Website Orders",
        overview["website_orders"]
    ])

    writer.writerow([
        "WhatsApp Orders",
        overview["whatsapp_orders"]
    ])

    writer.writerow([
        "Total Products",
        overview["total_products"]
    ])

    writer.writerow([
        "Average Orders / Day",
        overview["average_orders_per_day"]
    ])

    writer.writerow([])

    # --------------------------------------------------------
    # CATEGORY PERFORMANCE
    # --------------------------------------------------------

    writer.writerow(["CATEGORY PERFORMANCE"])

    writer.writerow([
        "Category",
        "Orders",
        "Order %"
    ])

    for item in analysis.get("category", []):

        writer.writerow([
            item.get("name", ""),
            item.get("count", 0),
            item.get("percentage", 0)
        ])

    writer.writerow([])

    # --------------------------------------------------------
    # MONTHLY CATEGORY PERFORMANCE
    # --------------------------------------------------------

    writer.writerow([
        "MONTHLY CATEGORY PERFORMANCE"
    ])

    writer.writerow([
        "Month",
        "Category",
        "Orders",
        "Growth %",
        "Growth Status"
    ])

    for month_data in analysis.get(
        "category_monthly",
        []
    ):

        month = month_data.get(
            "month",
            ""
        )

        for category_name, category_info in (
            month_data.get(
                "categories",
                {}
            ).items()
        ):

            writer.writerow([
                month,
                category_name,
                category_info.get(
                    "orders",
                    0
                ),
                category_info.get(
                    "growth_percentage",
                    0
                ),
                category_info.get(
                    "growth_status",
                    ""
                )
            ])

    writer.writerow([])

    # --------------------------------------------------------
    # PRODUCT INTELLIGENCE
    # --------------------------------------------------------

    writer.writerow([
        "PRODUCT INTELLIGENCE"
    ])

    writer.writerow([
        "Type",
        "Product ID",
        "Design Number",
        "Name",
        "Category",
        "Orders"
    ])

    for product in analysis.get(
        "product_intelligence",
        {}
    ).get(
        "best_sellers",
        []
    ):

        writer.writerow([
            "Best Seller",
            product.get("product_id", ""),
            product.get("design_number", ""),
            product.get("name", ""),
            product.get("category", ""),
            product.get("orders", 0)
        ])

    for product in analysis.get(
        "product_intelligence",
        {}
    ).get(
        "underperforming",
        []
    ):

        writer.writerow([
            "Underperforming",
            product.get("product_id", ""),
            product.get("design_number", ""),
            product.get("name", ""),
            product.get("category", ""),
            product.get("orders", 0)
        ])

    for product in analysis.get(
        "product_intelligence",
        {}
    ).get(
        "never_ordered",
        []
    ):

        writer.writerow([
            "Never Ordered",
            product.get("product_id", ""),
            product.get("design_number", ""),
            product.get("name", ""),
            product.get("category", ""),
            product.get("orders", 0)
        ])

    writer.writerow([])

    # --------------------------------------------------------
    # RETAILERS
    # --------------------------------------------------------

    writer.writerow([
        "RETAILER ANALYSIS"
    ])

    writer.writerow([
        "Retailer",
        "Total Orders",
        "Custom Orders",
        "Stock Orders"
    ])

    for retailer in analysis.get(
        "retailers",
        []
    ):

        writer.writerow([
            retailer.get(
                "retailer_name",
                ""
            ),
            retailer.get(
                "total_orders",
                0
            ),
            retailer.get(
                "custom_orders",
                0
            ),
            retailer.get(
                "stock_orders",
                0
            )
        ])

    writer.writerow([])

    # --------------------------------------------------------
    # METAL
    # --------------------------------------------------------

    writer.writerow([
        "METAL ANALYSIS"
    ])

    writer.writerow([
        "Metal",
        "Orders"
    ])

    for item in analysis.get(
        "metal",
        []
    ):

        writer.writerow([
            item.get("name", ""),
            item.get("count", 0)
        ])

    writer.writerow([])

    # --------------------------------------------------------
    # STONE
    # --------------------------------------------------------

    writer.writerow([
        "STONE ANALYSIS"
    ])

    writer.writerow([
        "Stone",
        "Orders"
    ])

    for item in analysis.get(
        "stone",
        []
    ):

        writer.writerow([
            item.get("name", ""),
            item.get("count", 0)
        ])

    writer.writerow([])

    # --------------------------------------------------------
    # STATUS
    # --------------------------------------------------------

    writer.writerow([
        "ORDER STATUS"
    ])

    writer.writerow([
        "Status",
        "Orders"
    ])

    for item in analysis.get(
        "status",
        []
    ):

        writer.writerow([
            item.get("name", ""),
            item.get("count", 0)
        ])

    writer.writerow([])

    # --------------------------------------------------------
    # DUE DATE
    # --------------------------------------------------------

    writer.writerow([
        "DUE DATE ANALYSIS"
    ])

    due_dates = analysis.get(
        "due_dates",
        {}
    )

    writer.writerow([
        "Due This Week",
        due_dates.get(
            "due_this_week",
            0
        )
    ])

    writer.writerow([
        "Due Next Week",
        due_dates.get(
            "due_next_week",
            0
        )
    ])

    writer.writerow([
        "Overdue",
        due_dates.get(
            "overdue",
            0
        )
    ])

    writer.writerow([
        "Completed On Time",
        due_dates.get(
            "completed_on_time",
            0
        )
    ])

    writer.writerow([
        "Delayed",
        due_dates.get(
            "delayed",
            0
        )
    ])

    writer.writerow([])

    # --------------------------------------------------------
    # AUTOMATIC INSIGHTS
    # --------------------------------------------------------

    writer.writerow([
        "AUTOMATIC BUSINESS INSIGHTS"
    ])

    writer.writerow([
        "Type",
        "Title",
        "Insight"
    ])

    for insight in analysis.get(
        "automatic_insights",
        []
    ):

        writer.writerow([
            insight.get("type", ""),
            insight.get("title", ""),
            insight.get("message", "")
        ])

    text_output.flush()

    # Detach the text wrapper so it doesn't close BytesIO
    text_output.detach()

    output.seek(0)

    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={
            "Content-Disposition":
                'attachment; filename="analysis_report.csv"'
        }
    )

# ============================================================
# ANALYSIS EXPORT — EXCEL
# ============================================================

@api_router.get("/admin/analysis/export/excel")
async def export_analysis_excel(
    request: Request,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    channel: Optional[str] = None,
    retailer_id: Optional[str] = None,
    category: Optional[str] = None,
    product_id: Optional[str] = None,
    order_type: Optional[str] = None,
    metal: Optional[str] = None,
    purity: Optional[str] = None,
    stone: Optional[str] = None,
):
    # Get the exact same filtered Analysis data
    analysis = await admin_combined_analysis(
        request=request,
        from_date=from_date,
        to_date=to_date,
        channel=channel,
        retailer_id=retailer_id,
        category=category,
        product_id=product_id,
        order_type=order_type,
        metal=metal,
        purity=purity,
        stone=stone,
    )

    wb = Workbook()

    # ========================================================
    # HELPER
    # ========================================================

    def setup_sheet(ws, title):

        ws.title = title

        ws.freeze_panes = "A2"

    def add_header(ws, row):

        for cell in row:

            cell.font = Font(
                bold=True
            )

    # ========================================================
    # 1. OVERVIEW
    # ========================================================

    ws = wb.active

    setup_sheet(
        ws,
        "Overview"
    )

    ws.append([
        "ANALYSIS REPORT"
    ])

    ws["A1"].font = Font(
        bold=True
    )

    ws.append([])

    ws.append([
        "FILTER",
        "VALUE"
    ])

    add_header(
        ws,
        ws[4]
    )

    filters = analysis.get(
        "filters",
        {}
    )

    filter_rows = [
        (
            "From Date",
            filters.get("from_date")
        ),
        (
            "To Date",
            filters.get("to_date")
        ),
        (
            "Channel",
            filters.get("channel")
        ),
        (
            "Retailer",
            filters.get("retailer_id")
        ),
        (
            "Category",
            filters.get("category")
        ),
        (
            "Product / Design",
            filters.get("product_id")
        ),
        (
            "Custom / Stock",
            filters.get("order_type")
        ),
        (
            "Metal",
            filters.get("metal")
        ),
        (
            "Purity",
            filters.get("purity")
        ),
        (
            "Stone",
            filters.get("stone")
        ),
    ]

    for row in filter_rows:

        ws.append([
            row[0],
            row[1]
        ])

    ws.append([])

    ws.append([
        "METRIC",
        "VALUE"
    ])

    add_header(
        ws,
        ws[17]
    )

    overview = analysis.get(
        "overview",
        {}
    )

    overview_rows = [
        (
            "Total Orders",
            overview.get(
                "total_orders",
                0
            )
        ),
        (
            "Combined Orders",
            overview.get(
                "combined_orders",
                0
            )
        ),
        (
            "Website Orders",
            overview.get(
                "website_orders",
                0
            )
        ),
        (
            "WhatsApp Orders",
            overview.get(
                "whatsapp_orders",
                0
            )
        ),
        (
            "Total Products",
            overview.get(
                "total_products",
                0
            )
        ),
        (
            "Average Orders / Day",
            overview.get(
                "average_orders_per_day",
                0
            )
        ),
    ]

    for row in overview_rows:

        ws.append([
            row[0],
            row[1]
        ])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 25

    # ========================================================
    # 2. CATEGORIES
    # ========================================================

    ws = wb.create_sheet(
        "Categories"
    )

    setup_sheet(
        ws,
        "Categories"
    )

    ws.append([
        "Category",
        "Orders",
        "Order %"
    ])

    add_header(
        ws,
        ws[1]
    )

    for item in analysis.get(
        "category",
        []
    ):

        ws.append([
            item.get(
                "name",
                ""
            ),
            item.get(
                "count",
                0
            ),
            item.get(
                "percentage",
                0
            )
        ])

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15

    # ========================================================
    # 3. MONTHLY PERFORMANCE
    # ========================================================

    ws = wb.create_sheet(
        "Monthly Performance"
    )

    setup_sheet(
        ws,
        "Monthly Performance"
    )

    ws.append([
        "Month",
        "Category",
        "Orders",
        "Growth %",
        "Growth Status"
    ])

    add_header(
        ws,
        ws[1]
    )

    for month_data in analysis.get(
        "category_monthly",
        []
    ):

        month = month_data.get(
            "month",
            ""
        )

        for category_name, category_info in (
            month_data.get(
                "categories",
                {}
            ).items()
        ):

            ws.append([
                month,
                category_name,
                category_info.get(
                    "orders",
                    0
                ),
                category_info.get(
                    "growth_percentage",
                    0
                ),
                category_info.get(
                    "growth_status",
                    ""
                )
            ])

    for column in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[column].width = 22

    # ========================================================
    # 4. PRODUCTS
    # ========================================================

    ws = wb.create_sheet(
        "Products"
    )

    setup_sheet(
        ws,
        "Products"
    )

    ws.append([
        "Type",
        "Product ID",
        "Design Number",
        "Name",
        "Category",
        "Orders"
    ])

    add_header(
        ws,
        ws[1]
    )

    intelligence = analysis.get(
        "product_intelligence",
        {}
    )

    for product in intelligence.get(
        "best_sellers",
        []
    ):

        ws.append([
            "Best Seller",
            product.get(
                "product_id",
                ""
            ),
            product.get(
                "design_number",
                ""
            ),
            product.get(
                "name",
                ""
            ),
            product.get(
                "category",
                ""
            ),
            product.get(
                "orders",
                0
            )
        ])

    for product in intelligence.get(
        "underperforming",
        []
    ):

        ws.append([
            "Underperforming",
            product.get(
                "product_id",
                ""
            ),
            product.get(
                "design_number",
                ""
            ),
            product.get(
                "name",
                ""
            ),
            product.get(
                "category",
                ""
            ),
            product.get(
                "orders",
                0
            )
        ])

    for product in intelligence.get(
        "never_ordered",
        []
    ):

        ws.append([
            "Never Ordered",
            product.get(
                "product_id",
                ""
            ),
            product.get(
                "design_number",
                ""
            ),
            product.get(
                "name",
                ""
            ),
            product.get(
                "category",
                ""
            ),
            product.get(
                "orders",
                0
            )
        ])

    for column in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[column].width = 24

    # ========================================================
    # 5. RETAILERS
    # ========================================================

    ws = wb.create_sheet(
        "Retailers"
    )

    setup_sheet(
        ws,
        "Retailers"
    )

    ws.append([
        "Retailer",
        "Total Orders",
        "Custom Orders",
        "Stock Orders"
    ])

    add_header(
        ws,
        ws[1]
    )

    for retailer in analysis.get(
        "retailers",
        []
    ):

        ws.append([
            retailer.get(
                "retailer_name",
                ""
            ),
            retailer.get(
                "total_orders",
                0
            ),
            retailer.get(
                "custom_orders",
                0
            ),
            retailer.get(
                "stock_orders",
                0
            )
        ])

    for column in ["A", "B", "C", "D"]:
        ws.column_dimensions[column].width = 25

    # ========================================================
    # 6. METAL
    # ========================================================

    ws = wb.create_sheet(
        "Metal"
    )

    setup_sheet(
        ws,
        "Metal"
    )

    ws.append([
        "Metal",
        "Orders"
    ])

    add_header(
        ws,
        ws[1]
    )

    for item in analysis.get(
        "metal",
        []
    ):

        ws.append([
            item.get(
                "name",
                ""
            ),
            item.get(
                "count",
                0
            )
        ])

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15

    # ========================================================
    # 7. STONE
    # ========================================================

    ws = wb.create_sheet(
        "Stone"
    )

    setup_sheet(
        ws,
        "Stone"
    )

    ws.append([
        "Stone",
        "Orders"
    ])

    add_header(
        ws,
        ws[1]
    )

    for item in analysis.get(
        "stone",
        []
    ):

        ws.append([
            item.get(
                "name",
                ""
            ),
            item.get(
                "count",
                0
            )
        ])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15

    # ========================================================
    # 8. STATUS
    # ========================================================

    ws = wb.create_sheet(
        "Status"
    )

    setup_sheet(
        ws,
        "Status"
    )

    ws.append([
        "Status",
        "Orders"
    ])

    add_header(
        ws,
        ws[1]
    )

    for item in analysis.get(
        "status",
        []
    ):

        ws.append([
            item.get(
                "name",
                ""
            ),
            item.get(
                "count",
                0
            )
        ])

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15

    # ========================================================
    # 9. DUE DATES
    # ========================================================

    ws = wb.create_sheet(
        "Due Dates"
    )

    setup_sheet(
        ws,
        "Due Dates"
    )

    ws.append([
        "Metric",
        "Orders"
    ])

    add_header(
        ws,
        ws[1]
    )

    due_dates = analysis.get(
        "due_dates",
        {}
    )

    due_rows = [
        (
            "Due This Week",
            due_dates.get(
                "due_this_week",
                0
            )
        ),
        (
            "Due Next Week",
            due_dates.get(
                "due_next_week",
                0
            )
        ),
        (
            "Overdue",
            due_dates.get(
                "overdue",
                0
            )
        ),
        (
            "Completed On Time",
            due_dates.get(
                "completed_on_time",
                0
            )
        ),
        (
            "Delayed",
            due_dates.get(
                "delayed",
                0
            )
        ),
    ]

    for row in due_rows:

        ws.append([
            row[0],
            row[1]
        ])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 15

    # ========================================================
    # 10. INSIGHTS
    # ========================================================

    ws = wb.create_sheet(
        "Business Insights"
    )

    setup_sheet(
        ws,
        "Business Insights"
    )

    ws.append([
        "Type",
        "Title",
        "Insight"
    ])

    add_header(
        ws,
        ws[1]
    )

    for insight in analysis.get(
        "automatic_insights",
        []
    ):

        ws.append([
            insight.get(
                "type",
                ""
            ),
            insight.get(
                "title",
                ""
            ),
            insight.get(
                "message",
                ""
            )
        ])

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 80

    # ========================================================
    # CREATE FILE
    # ========================================================

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
                'attachment; filename="analysis_report.xlsx"'
        }
    )


# ============================================================
# ANALYSIS EXPORT — PDF
# ============================================================

@api_router.get("/admin/analysis/export/pdf")
async def export_analysis_pdf(
    request: Request,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    channel: Optional[str] = None,
    retailer_id: Optional[str] = None,
    category: Optional[str] = None,
    product_id: Optional[str] = None,
    order_type: Optional[str] = None,
    metal: Optional[str] = None,
    purity: Optional[str] = None,
    stone: Optional[str] = None,
):
    # Get the exact same filtered Analysis data
    analysis = await admin_combined_analysis(
        request=request,
        from_date=from_date,
        to_date=to_date,
        channel=channel,
        retailer_id=retailer_id,
        category=category,
        product_id=product_id,
        order_type=order_type,
        metal=metal,
        purity=purity,
        stone=stone,
    )

    output = BytesIO()

    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()

    title_style = styles["Title"]
    heading_style = styles["Heading2"]
    normal_style = styles["BodyText"]

    story = []

    # ========================================================
    # HELPER FUNCTIONS
    # ========================================================

    def add_heading(text):

        story.append(
            Paragraph(
                str(text),
                heading_style
            )
        )

        story.append(
            Spacer(
                1,
                5
            )
        )

    def add_table(headers, rows):

        table_data = [
            headers
        ] + rows

        table = Table(
            table_data,
            repeatRows=1
        )

        table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.lightgrey
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold"
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP"
                ),
                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    8
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    5
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    4
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    4
                ),
            ])
        )

        story.append(table)

        story.append(
            Spacer(
                1,
                10
            )
        )

    # ========================================================
    # TITLE
    # ========================================================

    story.append(
        Paragraph(
            "Business Analysis Report",
            title_style
        )
    )

    story.append(
        Spacer(
            1,
            10
        )
    )

    # ========================================================
    # FILTERS
    # ========================================================

    add_heading(
        "Selected Filters"
    )

    filters = analysis.get(
        "filters",
        {}
    )

    filter_rows = [
        [
            "From Date",
            filters.get(
                "from_date"
            ) or "All"
        ],
        [
            "To Date",
            filters.get(
                "to_date"
            ) or "All"
        ],
        [
            "Channel",
            filters.get(
                "channel"
            ) or "All"
        ],
        [
            "Retailer",
            filters.get(
                "retailer_id"
            ) or "All"
        ],
        [
            "Category",
            filters.get(
                "category"
            ) or "All"
        ],
        [
            "Product / Design",
            filters.get(
                "product_id"
            ) or "All"
        ],
        [
            "Custom / Stock",
            filters.get(
                "order_type"
            ) or "All"
        ],
        [
            "Metal",
            filters.get(
                "metal"
            ) or "All"
        ],
        [
            "Purity",
            filters.get(
                "purity"
            ) or "All"
        ],
        [
            "Stone",
            filters.get(
                "stone"
            ) or "All"
        ],
    ]

    add_table(
        [
            "Filter",
            "Value"
        ],
        filter_rows
    )

    # ========================================================
    # OVERVIEW
    # ========================================================

    add_heading(
        "Overview"
    )

    overview = analysis.get(
        "overview",
        {}
    )

    overview_rows = [
        [
            "Total Orders",
            overview.get(
                "total_orders",
                0
            )
        ],
        [
            "Combined Orders",
            overview.get(
                "combined_orders",
                0
            )
        ],
        [
            "Website Orders",
            overview.get(
                "website_orders",
                0
            )
        ],
        [
            "WhatsApp Orders",
            overview.get(
                "whatsapp_orders",
                0
            )
        ],
        [
            "Total Products",
            overview.get(
                "total_products",
                0
            )
        ],
        [
            "Average Orders / Day",
            overview.get(
                "average_orders_per_day",
                0
            )
        ],
    ]

    add_table(
        [
            "Metric",
            "Value"
        ],
        overview_rows
    )

    # ========================================================
    # CATEGORY PERFORMANCE
    # ========================================================

    story.append(
        PageBreak()
    )

    add_heading(
        "Category Performance"
    )

    category_rows = []

    for item in analysis.get(
        "category",
        []
    ):

        category_rows.append([
            item.get(
                "name",
                ""
            ),
            item.get(
                "count",
                0
            ),
            item.get(
                "percentage",
                0
            )
        ])

    add_table(
        [
            "Category",
            "Orders",
            "Order %"
        ],
        category_rows
    )

    # ========================================================
    # MONTHLY CATEGORY GROWTH
    # ========================================================

    add_heading(
        "Monthly Category Performance"
    )

    monthly_rows = []

    for month_data in analysis.get(
        "category_monthly",
        []
    ):

        month = month_data.get(
            "month",
            ""
        )

        for category_name, category_info in (
            month_data.get(
                "categories",
                {}
            ).items()
        ):

            monthly_rows.append([
                month,
                category_name,
                category_info.get(
                    "orders",
                    0
                ),
                category_info.get(
                    "growth_percentage",
                    0
                ),
                category_info.get(
                    "growth_status",
                    ""
                )
            ])

    add_table(
        [
            "Month",
            "Category",
            "Orders",
            "Growth %",
            "Status"
        ],
        monthly_rows
    )

    # ========================================================
    # PRODUCT INTELLIGENCE
    # ========================================================

    story.append(
        PageBreak()
    )

    add_heading(
        "Product Catalogue Intelligence"
    )

    product_rows = []

    intelligence = analysis.get(
        "product_intelligence",
        {}
    )

    for product in intelligence.get(
        "best_sellers",
        []
    ):

        product_rows.append([
            "Best Seller",
            product.get(
                "product_id",
                ""
            ),
            product.get(
                "design_number",
                ""
            ),
            product.get(
                "name",
                ""
            ),
            product.get(
                "category",
                ""
            ),
            product.get(
                "orders",
                0
            )
        ])

    for product in intelligence.get(
        "underperforming",
        []
    ):

        product_rows.append([
            "Underperforming",
            product.get(
                "product_id",
                ""
            ),
            product.get(
                "design_number",
                ""
            ),
            product.get(
                "name",
                ""
            ),
            product.get(
                "category",
                ""
            ),
            product.get(
                "orders",
                0
            )
        ])

    for product in intelligence.get(
        "never_ordered",
        []
    ):

        product_rows.append([
            "Never Ordered",
            product.get(
                "product_id",
                ""
            ),
            product.get(
                "design_number",
                ""
            ),
            product.get(
                "name",
                ""
            ),
            product.get(
                "category",
                ""
            ),
            product.get(
                "orders",
                0
            )
        ])

    add_table(
        [
            "Type",
            "Product ID",
            "Design",
            "Name",
            "Category",
            "Orders"
        ],
        product_rows
    )

    # ========================================================
    # RETAILER ANALYSIS
    # ========================================================

    add_heading(
        "Retailer Analysis"
    )

    retailer_rows = []

    for retailer in analysis.get(
        "retailers",
        []
    ):

        retailer_rows.append([
            retailer.get(
                "retailer_name",
                ""
            ),
            retailer.get(
                "total_orders",
                0
            ),
            retailer.get(
                "custom_orders",
                0
            ),
            retailer.get(
                "stock_orders",
                0
            )
        ])

    add_table(
        [
            "Retailer",
            "Total Orders",
            "Custom",
            "Stock"
        ],
        retailer_rows
    )

    # ========================================================
    # METAL ANALYSIS
    # ========================================================

    add_heading(
        "Metal Analysis"
    )

    metal_rows = []

    for item in analysis.get(
        "metal",
        []
    ):

        metal_rows.append([
            item.get(
                "name",
                ""
            ),
            item.get(
                "count",
                0
            )
        ])

    add_table(
        [
            "Metal",
            "Orders"
        ],
        metal_rows
    )

    # ========================================================
    # STONE ANALYSIS
    # ========================================================

    add_heading(
        "Stone Analysis"
    )

    stone_rows = []

    for item in analysis.get(
        "stone",
        []
    ):

        stone_rows.append([
            item.get(
                "name",
                ""
            ),
            item.get(
                "count",
                0
            )
        ])

    add_table(
        [
            "Stone",
            "Orders"
        ],
        stone_rows
    )

    # ========================================================
    # STATUS
    # ========================================================

    add_heading(
        "Order Status"
    )

    status_rows = []

    for item in analysis.get(
        "status",
        []
    ):

        status_rows.append([
            item.get(
                "name",
                ""
            ),
            item.get(
                "count",
                0
            )
        ])

    add_table(
        [
            "Status",
            "Orders"
        ],
        status_rows
    )

    # ========================================================
    # DUE DATES
    # ========================================================

    add_heading(
        "Due Date Analysis"
    )

    due_dates = analysis.get(
        "due_dates",
        {}
    )

    due_rows = [
        [
            "Due This Week",
            due_dates.get(
                "due_this_week",
                0
            )
        ],
        [
            "Due Next Week",
            due_dates.get(
                "due_next_week",
                0
            )
        ],
        [
            "Overdue",
            due_dates.get(
                "overdue",
                0
            )
        ],
        [
            "Completed On Time",
            due_dates.get(
                "completed_on_time",
                0
            )
        ],
        [
            "Delayed",
            due_dates.get(
                "delayed",
                0
            )
        ],
    ]

    add_table(
        [
            "Metric",
            "Orders"
        ],
        due_rows
    )

    # ========================================================
    # AUTOMATIC INSIGHTS
    # ========================================================

    story.append(
        PageBreak()
    )

    add_heading(
        "Automatic Business Insights"
    )

    insight_rows = []

    for insight in analysis.get(
        "automatic_insights",
        []
    ):

        insight_rows.append([
            insight.get(
                "type",
                ""
            ),
            insight.get(
                "title",
                ""
            ),
            insight.get(
                "message",
                ""
            )
        ])

    add_table(
        [
            "Type",
            "Title",
            "Insight"
        ],
        insight_rows
    )

    # ========================================================
    # BUILD PDF
    # ========================================================

    doc.build(
        story
    )

    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
                'attachment; filename="analysis_report.pdf"'
        }
    )

@api_router.get("/admin/whatsapp-orders/{order_id}")
async def admin_get_whatsapp_order(
    order_id: str,
    request: Request
):
    await get_admin_user(request)

    order = await whatsapp_orders.find_one(
    {"orderId": order_id},
    {"_id": 0}
)

    if not order:
        raise HTTPException(
            status_code=404,
            detail="Order not found"
        )

    return order

@api_router.get("/admin/whatsapp-orders/{order_id}/excel")
async def download_whatsapp_order_excel(
    order_id: str,
    request: Request
):
    await get_admin_user(request)

    order = await whatsapp_orders.find_one(
        {"orderId": order_id},
        {"_id": 0}
    )

    if not order:
        raise HTTPException(
            status_code=404,
            detail="Order not found"
        )

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Order Details"

    # Header
    ws["A1"] = "Field"
    ws["B1"] = "Value"

    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    # Fields to exclude
    excluded_fields = {
        "design_images",
        "reference_video",
        "flow_token"
    }

    row = 2

    for key, value in order.items():
        if key in excluded_fields:
            continue

        ws.cell(row=row, column=1).value = key.replace("_", " ").title()
        ws.cell(row=row, column=2).value = str(value) if value is not None else ""
        row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{order_id}.xlsx"'
        },
    )
@api_router.get("/admin/whatsapp-orders/excel/today")
async def download_today_orders_excel(request: Request):
    await get_admin_user(request)

    from datetime import datetime

    today = datetime.now().strftime("%Y-%m-%d")

    orders = await whatsapp_orders.find(
        {"order_date": today},
        {"_id": 0}
    ).to_list(None)

    wb = Workbook()
    ws = wb.active
    ws.title = "Today's Orders"

    excluded_fields = {
        "design_images",
        "reference_video",
        "flow_token"
    }

    if orders:

        headers = [
            key
            for key in orders[0].keys()
            if key not in excluded_fields
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header.replace("_", " ").title()
            cell.font = Font(bold=True)

        for row_index, order in enumerate(orders, start=2):

            for col, header in enumerate(headers, start=1):

                ws.cell(
                    row=row_index,
                    column=col
                ).value = str(order.get(header, ""))

    output = BytesIO()

    wb.save(output)

    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            f'attachment; filename="Today_Orders_{today}.xlsx"'
        },
    )

@api_router.get("/admin/whatsapp-orders/excel/date/{selected_date}")
async def download_orders_by_date(
    selected_date: str,
    request: Request
):
    await get_admin_user(request)

    orders = await whatsapp_orders.find(
        {"order_date": selected_date},
        {"_id": 0}
    ).to_list(length=None)

    if not orders:
        raise HTTPException(
            status_code=404,
            detail="No orders found for this date."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = f"Orders {selected_date}"

    excluded_fields = {
        "design_images",
        "reference_video",
        "flow_token"
    }

    headers = [
        key for key in orders[0].keys()
        if key not in excluded_fields
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header.replace("_", " ").title()
        cell.font = Font(bold=True)

    for row_index, order in enumerate(orders, start=2):
        for col, header in enumerate(headers, start=1):
            ws.cell(
                row=row_index,
                column=col
            ).value = str(order.get(header, ""))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            f'attachment; filename="Orders_{selected_date}.xlsx"'
        },
    )


@api_router.get("/admin/whatsapp-orders/excel/customer/{customer_name}")
async def download_orders_by_customer(
    customer_name: str,
    request: Request
):
    await get_admin_user(request)

    orders = await whatsapp_orders.find(
        {
            "customer_name": {
                "$regex": f"^{customer_name}$",
                "$options": "i"
            }
        },
        {"_id": 0}
    ).to_list(length=None)

    if not orders:
        raise HTTPException(
            status_code=404,
            detail="No orders found for this customer."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = customer_name

    excluded_fields = {
        "design_images",
        "reference_video",
        "flow_token"
    }

    headers = [
        key for key in orders[0].keys()
        if key not in excluded_fields
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header.replace("_", " ").title()
        cell.font = Font(bold=True)

    for row_index, order in enumerate(orders, start=2):
        for col, header in enumerate(headers, start=1):
            ws.cell(
                row=row_index,
                column=col
            ).value = str(order.get(header, ""))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            f'attachment; filename="{customer_name}_Orders.xlsx"'
        },
    )


@api_router.get("/admin/whatsapp-orders/excel/customer-date")
async def download_orders_by_customer_and_date(
    customer_name: str,
    order_date: str,
    request: Request
):
    await get_admin_user(request)

    orders = await whatsapp_orders.find(
        {
            "customer_name": {
                "$regex": f"^{customer_name}$",
                "$options": "i"
            },
            "order_date": order_date
        },
        {"_id": 0}
    ).to_list(length=None)

    if not orders:
        raise HTTPException(
            status_code=404,
            detail="No matching orders found."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = customer_name[:31]

    excluded_fields = {
        "design_images",
        "reference_video",
        "flow_token"
    }

    headers = [
        key
        for key in orders[0].keys()
        if key not in excluded_fields
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header.replace("_", " ").title()
        cell.font = Font(bold=True)

    for row_index, order in enumerate(orders, start=2):
        for col, header in enumerate(headers, start=1):
            ws.cell(
                row=row_index,
                column=col
            ).value = str(order.get(header, ""))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            f'attachment; filename="{customer_name}_{order_date}.xlsx"'
        },
    )

@api_router.get("/admin/whatsapp-orders/excel/date-range")
async def download_orders_by_date_range(
    from_date: str,
    to_date: str,
    request: Request
):
    await get_admin_user(request)

    orders = await whatsapp_orders.find(
        {
            "order_date": {
                "$gte": from_date,
                "$lte": to_date
            }
        },
        {"_id": 0}
    ).sort("order_date", 1).to_list(length=None)

    if not orders:
        raise HTTPException(
            status_code=404,
            detail="No orders found for the selected date range."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    excluded_fields = {
        "design_images",
        "reference_video",
        "flow_token"
    }

    headers = [
        key
        for key in orders[0].keys()
        if key not in excluded_fields
    ]

    # Header row
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header.replace("_", " ").title()
        cell.font = Font(bold=True)

    # Order rows
    for row_index, order in enumerate(orders, start=2):
        for col, header in enumerate(headers, start=1):
            ws.cell(
                row=row_index,
                column=col
            ).value = str(order.get(header, ""))

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            f'attachment; filename="Orders_{from_date}_to_{to_date}.xlsx"'
        },
    )

@api_router.get("/admin/whatsapp-orders/excel/customers-date")
async def download_orders_by_customers_and_date(
    request: Request,
    customer_names: List[str] = Query(...),
    order_date: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
):
    await get_admin_user(request)

    # Validate customers
    if not customer_names:
        raise HTTPException(
            status_code=400,
            detail="Please select at least one customer."
        )

    # Clean customer names
    cleaned_customer_names = []

    for name in customer_names:
        if not name:
            continue

        for customer in name.split(","):
            customer = customer.strip()

            if customer:
                cleaned_customer_names.append(customer)

    customer_names = list(dict.fromkeys(cleaned_customer_names))

    if not customer_names:
        raise HTTPException(
            status_code=400,
            detail="Please select at least one customer."
        )

    # Validate date selection
    if order_date:
        date_query = {
            "order_date": order_date
        }

    elif from_date and to_date:
        if from_date > to_date:
            raise HTTPException(
                status_code=400,
                detail="From date cannot be after To date."
            )

        date_query = {
            "order_date": {
                "$gte": from_date,
                "$lte": to_date
            }
        }

    else:
        raise HTTPException(
            status_code=400,
            detail="Please select a date or a date range."
        )

    # Match selected customers
    customer_query = {
        "$or": [
            {
                "customer_name": {
                    "$regex": f"^\\s*{re.escape(name)}\\s*$",
                    "$options": "i"
                }
            }
            for name in customer_names
        ]
    }

    # Combine customer + date filters
    query = {
        **customer_query,
        **date_query
    }

    # Get matching orders
    orders = await whatsapp_orders.find(
        query,
        {"_id": 0}
    ).sort(
        "order_date", 1
    ).to_list(length=None)

    if not orders:
        raise HTTPException(
            status_code=404,
            detail="No matching orders found."
        )

    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    excluded_fields = {
        "design_images",
        "reference_video",
        "flow_token"
    }

    headers = [
        key
        for key in orders[0].keys()
        if key not in excluded_fields
    ]

    # Header row
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header.replace("_", " ").title()
        cell.font = Font(bold=True)

    # Order rows
    for row_index, order in enumerate(orders, start=2):
        for col, header in enumerate(headers, start=1):
            ws.cell(
                row=row_index,
                column=col
            ).value = str(order.get(header, ""))

    # Create Excel file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Filename
    if order_date:
        filename = f"Orders_{order_date}.xlsx"
    else:
        filename = f"Orders_{from_date}_to_{to_date}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
    )


# ===============================
# GET ALL CUSTOMER NAMES (NEW API)
# ===============================

@app.get("/admin/whatsapp-orders/customers")
async def get_all_customers():
    try:
        customers = await whatsapp_orders.distinct("customer_name")
        customers = [c for c in customers if c]

        return {
            "success": True,
            "customers": customers
        }

    except Exception as e:
        print("Error fetching customers:", str(e))
        return {
            "success": False,
            "customers": []
        }
# @api_router.put("/admin/whatsapp-orders/{order_id}")
# async def update_whatsapp_order(
#     order_id: str,
#     request: Request,
#     data: dict
# ):
#     await get_admin_user(request)

#     result = await whatsapp_orders.update_one(
#         {"orderId": order_id},
#         {
#             "$set": {
#                 "status": data.get("status"),
#                 "priority": data.get("priority"),
#                 "assignedTo": data.get("assignedTo"),
#                 "adminNotes": data.get("adminNotes"),
#             }
#         }
#     )

#     if result.matched_count == 0:
#         raise HTTPException(status_code=404, detail="Order not found")

#     order = await whatsapp_orders.find_one(
#         {"orderId": order_id},
#         {"_id": 0}
#     )

#     return order


@api_router.put("/admin/whatsapp-orders/{order_id}")
async def admin_update_whatsapp_order(
    order_id: str,
    update: WhatsAppOrderUpdate,
    request: Request
):
    await get_admin_user(request)

    update_data = {
        key: value
        for key, value in update.model_dump().items()
        if value is not None
    }

    if not update_data:
        raise HTTPException(
            status_code=400,
            detail="No data to update."
        )
    
        # Store the actual completion time when an order
    # is moved to a completed/delivered status.
    new_status = update_data.get("status")

    if new_status:
        normalized_status = str(new_status).strip().lower()

        if normalized_status in {"completed", "delivered"}:
            update_data["completedAt"] = datetime.now(timezone.utc)

        update_data["updatedAt"] = datetime.now(timezone.utc)

    result = await whatsapp_orders.update_one(
        {"orderId": order_id},
        {
            "$set": update_data
        }
    )

    if result.matched_count == 0:
        raise HTTPException(
            status_code=404,
            detail="Order not found."
        )

    return {
        "message": "Order updated successfully."
    }

from cloudinary.uploader import destroy

@api_router.delete("/admin/whatsapp-orders/{order_id}")
async def admin_delete_whatsapp_order(
    order_id: str,
    request: Request
):
    await get_admin_user(request)

    order = await whatsapp_orders.find_one(
        {"orderId": order_id}
    )

    if not order:
        raise HTTPException(
            status_code=404,
            detail="Order not found"
        )

    # Delete design images
    for image in order.get("design_images", []):

        try:

            public_id = image.split("/upload/")[1]
            public_id = public_id.split("/", 1)[1]
            public_id = public_id.rsplit(".", 1)[0]

            print("Deleting Cloudinary image:", public_id)

            destroy(public_id)

        except Exception as e:
            print("Image delete failed:", e)

    # Delete reference video
    if order.get("reference_video"):

        try:

            public_id = order["reference_video"].split("/upload/")[1]
            public_id = public_id.split("/", 1)[1]
            public_id = public_id.rsplit(".", 1)[0]

            print("Deleting Cloudinary video:", public_id)

            destroy(
                public_id,
                resource_type="video"
            )

        except Exception as e:
            print("Video delete failed:", e)

    await whatsapp_orders.delete_one(
        {"orderId": order_id}
    )

    return {
        "success": True
    }



# ──── SEEDING ────
async def seed_admin():
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@example.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")

    existing = await db.users.find_one({"email": admin_email})

    if existing is None:
        await db.users.insert_one({
            "email": admin_email,
            "password_hash": hash_password(admin_password),
            "name": "Admin",
            "role": "admin",
            "approved": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"Admin seeded: {admin_email}")

    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one(
            {"email": admin_email},
            {"$set": {"password_hash": hash_password(admin_password)}}
        )
        logger.info("Admin password updated")

    


    # ✅ SAFE MEMORY DIRECTORY FIX
    try:
        os.makedirs("memory", exist_ok=True)

        with open("memory/test_credentials.md", "w") as f:
            f.write(f"# Test Credentials\n\n")
            f.write(f"## Admin\n- Email: {admin_email}\n- Password: {admin_password}\n- Role: admin\n\n")
            f.write("## Auth Endpoints\n")
            f.write("- POST /api/auth/register\n")
            f.write("- POST /api/auth/login\n")
            f.write("- GET /api/auth/me\n")

    except Exception as e:
        logger.warning(f"Memory folder skipped: {e}")
        

async def seed_products():
    count = await db.products.count_documents({})
    if count > 0:
        logger.info(f"Products already exist ({count})")
        return
    products = []
    for cat in CATEGORIES:
        images = STOCK_IMAGES.get(cat["name"], ["https://images.unsplash.com/photo-1587947330318-88fcd9055420?w=500"])
        for i in range(30):
            pid = f"{cat['prefix']}-{random.randint(100000, 999999)}"
            img = images[i % len(images)]
            products.append({
                "product_id": pid, "category": cat["name"], "category_slug": cat["slug"],
                "images": [img, img, img], "rating": 5.0,
                "created_at": datetime.now(timezone.utc).isoformat()
            })
    await db.products.insert_many(products)
    logger.info(f"Seeded {len(products)} products")

@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    logger.info("AWS S3 storage initialized")
    await seed_admin()
    await seed_products()

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[os.environ.get("FRONTEND_URL", "http://localhost:3000"), "http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown():
    client.close()

@app.on_event("shutdown")
async def shutdown():
    client.close()


@app.get("/webhook")
async def verify_webhook(request: Request):

    mode = request.query_params.get("hub.mode")
    token = request.query_params.get("hub.verify_token")
    challenge = request.query_params.get("hub.challenge")

    if (
        mode == "subscribe"
        and token == os.getenv("VERIFY_TOKEN")
    ):
        print("✅ Webhook Verified")
        return PlainTextResponse(content=challenge)

    return PlainTextResponse(
        content="Verification Failed",
        status_code=403
    )


@app.post("/webhook")
async def whatsapp_webhook(request: Request):

    body = await request.json()

    print("========== NEW WEBHOOK ==========")
    print(body)
    print("=================================")

    try:


        value = (
            body.get("entry", [{}])[0]
            .get("changes", [{}])[0]
            .get("value", {})
        )

        print("VALUE KEYS:", list(value.keys()))
        print(value)

        messages = value.get("messages", [])

        print("MESSAGES:", messages)

        if not messages:
            return {"success": True}

        message = messages[0]

        if message:

            print("Message Type:", message.get("type"))

            # ==========================
            # VIDEO MESSAGE
            # ==========================

            if message.get("type") == "video":

                print("========== VIDEO RECEIVED ==========")

                sender = message["from"]

                if sender not in pending_video_uploads:
                    send_text_message(
                        sender,
                        "No pending order found for this video."
                    )
                    return {"success": True}

                order_id = pending_video_uploads[sender]

                video_id = message["video"]["id"]

                print("Uploading video:", video_id)

                uploaded = upload_whatsapp_video(video_id)

                result = await whatsapp_orders.update_one(
                    {"orderId": order_id},
                    {
                        "$set": {
                            "reference_video": uploaded["secure_url"]
                        }
                    }
                )

                print("ORDER ID:", order_id)
                print("VIDEO URL:", uploaded["secure_url"])
                print("MATCHED:", result.matched_count)
                print("MODIFIED:", result.modified_count)

                doc = await whatsapp_orders.find_one(
                    {"orderId": order_id},
                    {"_id": 0}
                )

                print("UPDATED DOC:")
                print(doc)

                                # ========================================================
                # REFERENCE VIDEO SUCCESSFULLY RECEIVED
                # ========================================================

                send_text_message(
                    sender,
                    "✅ Reference video received successfully."
                )

                # --------------------------------------------------------
                # Cancel the video timer
                # --------------------------------------------------------

                timer = video_upload_timers.pop(sender, None)

                if timer:
                    timer.cancel()

                video_waiting_users.pop(sender, None)

                # --------------------------------------------------------
                # GET UPDATED ORDER
                # --------------------------------------------------------

                order = await whatsapp_orders.find_one(
                    {
                        "orderId": order_id
                    },
                    {
                        "_id": 0
                    }
                )

                if not order:

                    print(
                        "Could not find order for PDF:",
                        order_id
                    )

                    pending_video_uploads.pop(sender, None)

                    return {"success": True}

                # --------------------------------------------------------
                # CREATE PDF
                # --------------------------------------------------------

                try:

                    pdf_path = create_order_pdf(order)

                    print(
                        "PDF generated successfully:",
                        pdf_path
                    )

                except Exception as e:

                    print(
                        "PDF generation failed:",
                        str(e)
                    )

                    send_text_message(
                        sender,
                        """⚠️ Reference video received successfully.

However, we could not generate your order PDF."""
                    )

                    pending_video_uploads.pop(sender, None)

                    return {"success": True}

                # --------------------------------------------------------
                # UPLOAD PDF TO S3
                # --------------------------------------------------------

                try:

                    with open(pdf_path, "rb") as pdf_file:

                        pdf_url = upload_to_s3(
                            pdf_file,
                            f"orders/pdfs/{order_id}.pdf",
                            "application/pdf"
                        )

                    print(
                        "PDF uploaded successfully:",
                        pdf_url
                    )

                except Exception as e:

                    print(
                        "PDF S3 upload failed:",
                        str(e)
                    )

                    send_text_message(
                        sender,
                        """⚠️ Reference video received successfully.

However, we could not send your order PDF."""
                    )

                    pending_video_uploads.pop(sender, None)

                    return {"success": True}

                # --------------------------------------------------------
                # SEND PDF TO CUSTOMER
                # --------------------------------------------------------

                send_document(
                    sender,
                    pdf_url,
                    f"{order_id}.pdf",
                    "📄 Your Jewellery Manufacturing Order"
                )

                # --------------------------------------------------------
                # CLEAR PENDING ORDER
                # --------------------------------------------------------

                pending_video_uploads.pop(sender, None)

                print(
                    f"Order PDF sent successfully for {order_id}"
                )

                return {"success": True}

            if message.get("type") == "text":

                text = (
                    message.get("text", {})
                    .get("body", "")
                    .lower()
                    .strip()
                )

                sender = message["from"]

                print("Message:", text)

                # Start Flow
                if text == "hi":
                    send_flow(sender)
                    return {"success": True}

                # Customer wants to upload video
                if text == "yes":

                    if sender not in pending_video_uploads:
                        send_text_message(
                            sender,
                            "No recent order found."
                        )
                        return {"success": True}

                    # from datetime import datetime, timedelta

                    video_waiting_users[sender] = (
                        datetime.utcnow() + timedelta(minutes=10)
                    )

                    timer = threading.Timer(
                        VIDEO_TIMEOUT,
                        video_upload_timeout,
                        args=[sender]
                    )

                    timer.start()
                    video_upload_timers[sender] = timer

                    # # START TIMER
                    # timer = threading.Timer(
                    #     VIDEO_TIMEOUT,
                    #     video_upload_timeout,
                    #     args=[sender]
                    # )

                    # timer.start()
                    # video_upload_timers[sender] = timer

                    send_text_message(
                        sender,
                        """🎥 Great!

            Please upload ONE reference video.
            Maximum size: 40 MB.

            You have 10 minutes."""
                    )

                    return {"success": True}

                # Customer doesn't want to upload video
                                # Customer doesn't want to upload video
                if text == "no":

                    # ========================================================
                    # CUSTOMER DOES NOT WANT TO UPLOAD VIDEO
                    # ========================================================

                    # IMPORTANT:
                    # Get the order ID BEFORE removing it from memory.
                    order_id = pending_video_uploads.get(sender)

                    if not order_id:
                        send_text_message(
                            sender,
                            """⚠️ We could not find your order.

Please contact our team."""
                        )

                        return {"success": True}

                    # --------------------------------------------------------
                    # Cancel the 10-minute video timer
                    # --------------------------------------------------------

                    timer = video_upload_timers.pop(sender, None)

                    if timer:
                        timer.cancel()

                    # --------------------------------------------------------
                    # Stop waiting for video
                    # --------------------------------------------------------

                    video_waiting_users.pop(sender, None)

                    # --------------------------------------------------------
                    # Get the complete order from MongoDB
                    # --------------------------------------------------------

                    order = await whatsapp_orders.find_one(
                        {
                            "orderId": order_id
                        },
                        {
                            "_id": 0
                        }
                    )

                    if not order:
                        send_text_message(
                            sender,
                            """⚠️ We could not find your order.

Please contact our team."""
                        )

                        pending_video_uploads.pop(sender, None)

                        return {"success": True}

                    # --------------------------------------------------------
                    # CREATE PDF
                    # --------------------------------------------------------

                    try:

                        pdf_path = create_order_pdf(order)

                        print(
                            "PDF generated successfully:",
                            pdf_path
                        )

                    except Exception as e:

                        print(
                            "PDF generation failed:",
                            str(e)
                        )

                        send_text_message(
                            sender,
                            """⚠️ Your order was submitted successfully, but we could not generate the PDF.

Our team will contact you."""
                        )

                        pending_video_uploads.pop(sender, None)

                        return {"success": True}

                    # --------------------------------------------------------
                    # UPLOAD PDF TO S3
                    # --------------------------------------------------------

                    try:

                        with open(pdf_path, "rb") as pdf_file:

                            pdf_url = upload_to_s3(
                                pdf_file,
                                f"orders/pdfs/{order_id}.pdf",
                                "application/pdf"
                            )

                        print(
                            "PDF uploaded successfully:",
                            pdf_url
                        )

                    except Exception as e:

                        print(
                            "PDF S3 upload failed:",
                            str(e)
                        )

                        send_text_message(
                            sender,
                            """⚠️ Your order was submitted successfully, but we could not send the PDF.

Our team will contact you."""
                        )

                        pending_video_uploads.pop(sender, None)

                        return {"success": True}

                    # --------------------------------------------------------
                    # SEND PDF TO CUSTOMER
                    # --------------------------------------------------------

                    send_document(
                        sender,
                        pdf_url,
                        f"{order_id}.pdf",
                        "📄 Your Jewellery Manufacturing Order"
                    )

                    # --------------------------------------------------------
                    # CLEAR PENDING ORDER
                    # --------------------------------------------------------

                    pending_video_uploads.pop(sender, None)

                    # --------------------------------------------------------
                    # FINAL MESSAGE
                    # --------------------------------------------------------

                    send_text_message(
                        sender,
                        """✅ Your order PDF has been sent successfully."""
                    )

                    return {"success": True}

    # --------------------------------------------------------
    # SEND PDF
    # --------------------------------------------------------

    # TEMPORARY:
    # We will connect this to a public PDF URL in the next step.

                    send_text_message(
                        sender,
                        """✅ Thank you!

                Your order has been submitted successfully."""
                    )

                    return {"success": True}
                    

                    
                    

                

            elif (
                message.get("type") == "interactive"
                and message["interactive"].get("type") == "nfm_reply"
            ):

                import json

                form_data = json.loads(
                    message["interactive"]["nfm_reply"]["response_json"]
                )

                print("========== FORM DATA ==========")

                for key, value in form_data.items():
                    print(f"{key}: {value}")

                print("===============================")


                cloudinary_images = []

                for image in form_data.get("design_images", []):

                    print("Uploading:", image["id"])

                    uploaded = upload_whatsapp_image(image["id"])

                    cloudinary_images.append(uploaded["secure_url"])

                order = form_data.copy()
                order["design_images"] = cloudinary_images
                order["orderId"] = await get_next_order_id()
                order["status"] = "New"
                order["priority"] = "Normal"
                order["assignedTo"] = ""
                order["adminNotes"] = ""
                order["createdAt"] = datetime.utcnow()

                print("BEFORE INSERT")

                result = await whatsapp_orders.insert_one(order)

                print("AFTER INSERT")
                print(result.inserted_id)

                count = await whatsapp_orders.count_documents({})
                print("TOTAL ORDERS:", count)

                pending_video_uploads[message["from"]] = order["orderId"]
                # Remember this customer's latest order so the next video can be attached
                send_text_message(
                    message["from"],
                    f"""✅ Thank you for your order!

                Your Order ID: {order["orderId"]}

                Would you like to upload a reference video?

                Reply with:

                YES
                or
                NO"""
                )

                # print("Send message response:", response)

        return {"success": True}

    except Exception as e:
        print(e)
        return {"success": False}
    

@app.post("/whatsapp/flow")
async def whatsapp_flow_endpoint(request: Request):

    body = await request.json()

    print("========== ENCRYPTED FLOW ==========")
    print(body)
    print("====================================")

    try:
        print("STEP 1")

        data, aes_key, iv = decrypt_request(body)

        print("STEP 2")

        print("Decrypted request:")
        print(data)

        action = data.get("action")

        if action == "ping":

            response = {
                "version": "3.0",
                 "data": {
                    "status": "active"
                }
            }
            
        elif action == "INIT":

            response = {
                "version": "3.0",
                "screen": "JEWELLERY_ORDER",
                "data": {}
             }
            

            action = data.get("action")

        if action == "ping":

            response = {
                "version": "3.0",
                "data": {
                    "status": "active"
                }
            }

        elif action == "INIT":

            response = {
                "version": "3.0",
                "screen": "JEWELLERY_ORDER",
                "data": {}
            }

        elif action == "BACK":

            response = {
                "version": "3.0",
                "screen": data.get("screen"),
                "data": {}
            }

        elif action == "data_exchange":

            response = {
                "version": "3.0",
                "screen": "SUCCESS",
                "data": {}
            }

        else:

            response = {
                "version": "3.0",
                "data": {
                    "status": "active"
                }
            }

        encrypted = encrypt_response(response, aes_key, iv)

        return Response(
            content=encrypted,
            media_type="text/plain"
        )
    
    except Exception as e:
        import traceback

        print("FLOW ERROR")
        traceback.print_exc()

        return {"error": str(e)}